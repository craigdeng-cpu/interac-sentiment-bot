"""
Interac Sentiment Analysis Bot
- Scrapes Reddit, X, RedFlagDeals, news for Interac mentions 4x/day
- Splits people vs press signals
- Alerts on sentiment drops
- Configurable via prompts.json
"""

import os
import json
import logging
import smtplib
import asyncio
import re
import html
from time import monotonic
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime, timezone, timedelta
from pathlib import Path
from collections import defaultdict
from urllib.parse import parse_qs, urlparse, quote_plus

import httpx
from ddgs import DDGS
from selenium import webdriver
from selenium.common.exceptions import WebDriverException, TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from telegram import Update
from telegram.error import BadRequest
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ─── Config ───────────────────────────────────────────────────────────────────
TELEGRAM_TOKEN = os.environ["TELEGRAM_TOKEN"]
KIMI_API_KEY = os.environ["KIMI_API_KEY"]
KIMI_API_URL = os.environ.get("KIMI_API_URL", "https://api.moonshot.ai/v1/chat/completions")
KIMI_MODEL = os.environ.get("KIMI_MODEL", "kimi-k2.5-preview")
PORT = int(os.environ.get("PORT", 3978))
WEBHOOK_URL = os.environ.get("WEBHOOK_URL", "")
ADMIN_IDS = {int(x) for x in os.environ.get("ADMIN_IDS", "").split(",") if x}
DAILY_LIMIT = int(os.environ.get("DAILY_LIMIT", "5"))

EMAIL_ENABLED = os.environ.get("EMAIL_ENABLED", "0") == "1"
EMAIL_SEND_MODE = os.environ.get("EMAIL_SEND_MODE", "alert").lower()
EMAIL_ALERT_DEDUP = os.environ.get("EMAIL_ALERT_DEDUP", "1") == "1"
EMAIL_COOLDOWN_MINUTES = int(os.environ.get("EMAIL_COOLDOWN_MINUTES", "0"))
EMAIL_WEEKLY_DAY = os.environ.get("EMAIL_WEEKLY_DAY", "monday").strip().lower()
EMAIL_WEEKLY_HOUR = int(os.environ.get("EMAIL_WEEKLY_HOUR", "9"))
EMAIL_PROVIDER = os.environ.get("EMAIL_PROVIDER", "smtp").strip().lower()

SMTP_HOST = os.environ.get("SMTP_HOST", "")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USERNAME = os.environ.get("SMTP_USERNAME", "")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")
RESEND_API_KEY = os.environ.get("RESEND_API_KEY", "")
RESEND_API_URL = os.environ.get("RESEND_API_URL", "https://api.resend.com/emails")
EMAIL_FROM = os.environ.get("EMAIL_FROM", "")
EMAIL_TO = [x.strip() for x in os.environ.get("EMAIL_TO", "").split(",") if x.strip()]
EMAIL_SUBJECT_PREFIX = os.environ.get("EMAIL_SUBJECT_PREFIX", "Interac Intelligence")
SCRAPE_TIMEOUT_SECONDS = int(os.environ.get("SCRAPE_TIMEOUT_SECONDS", "20"))
SCRAPE_MAX_PAGES_PER_SOURCE = int(os.environ.get("SCRAPE_MAX_PAGES_PER_SOURCE", "4"))
SCRAPE_MAX_RESULTS_PER_QUERY = int(os.environ.get("SCRAPE_MAX_RESULTS_PER_QUERY", "5"))
SCRAPE_USER_AGENT = os.environ.get(
    "SCRAPE_USER_AGENT",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
)
CHROMIUM_BINARY = os.environ.get("CHROMIUM_BINARY", "/usr/bin/chromium")
CHROMEDRIVER_PATH = os.environ.get("CHROMEDRIVER_PATH", "/usr/bin/chromedriver")
# Biweekly scan + /scan + /email (override on Railway if Selenium still hits the cap).
BIWEEKLY_FETCH_TIMEOUT = int(os.environ.get("BIWEEKLY_FETCH_TIMEOUT", "360"))
BIWEEKLY_ANALYZE_TIMEOUT = int(os.environ.get("BIWEEKLY_ANALYZE_TIMEOUT", "120"))
FETCH_FALLBACK_DDG = os.environ.get("FETCH_FALLBACK_DDG", "1") == "1"
SCRAPE_MAX_CONCURRENT_BROWSERS = max(1, int(os.environ.get("SCRAPE_MAX_CONCURRENT_BROWSERS", "2")))

EST = timezone(timedelta(hours=-5))

NO_MENTIONS_MARKER = "No mentions found in the past month."
LAST_FETCH_DIAGNOSTICS: dict = {}
_selenium_browser_sem = asyncio.Semaphore(SCRAPE_MAX_CONCURRENT_BROWSERS)

subscribed_chats: set[int] = set()
last_report: str = ""
last_mentions_raw: str = ""
last_email_sent_at: datetime | None = None
last_weekly_email_key: str | None = None

# Per-user daily rate limiting
user_usage: dict[int, dict] = defaultdict(lambda: {"count": 0, "date": None})
active_tasks: set[asyncio.Task] = set()


def _track_current_task() -> asyncio.Task | None:
    task = asyncio.current_task()
    if task is not None:
        active_tasks.add(task)
    return task


def _untrack_task(task: asyncio.Task | None) -> None:
    if task is not None:
        active_tasks.discard(task)


def _cancel_active_tasks(*, exclude: asyncio.Task | None = None) -> int:
    cancelled = 0
    for task in list(active_tasks):
        if exclude is not None and task is exclude:
            continue
        if task.done():
            active_tasks.discard(task)
            continue
        task.cancel()
        cancelled += 1
    return cancelled


def now_est() -> str:
    return datetime.now(EST).strftime("%Y-%m-%d %I:%M %p EST")


def check_rate_limit(user_id: int) -> tuple[bool, int]:
    if user_id in ADMIN_IDS:
        return True, -1

    today = datetime.now(EST).date()
    usage = user_usage[user_id]

    if usage["date"] != today:
        usage["count"] = 0
        usage["date"] = today

    if usage["count"] >= DAILY_LIMIT:
        return False, 0

    usage["count"] += 1
    return True, DAILY_LIMIT - usage["count"]


# ─── Prompt Config ────────────────────────────────────────────────────────────
def load_prompts() -> dict:
    base_dir = Path(__file__).parent
    config_path = base_dir / "prompts.json"
    with open(config_path) as f:
        config = json.load(f)

    prompt_files = config.get("prompt_files", {})
    default_prompt_files = {
        "followup_prompt": "prompts/followup_prompt.md",
    }
    for prompt_key, default_path in default_prompt_files.items():
        rel_path = prompt_files.get(prompt_key, default_path)
        prompt_path = base_dir / rel_path
        if prompt_path.exists():
            config[prompt_key] = prompt_path.read_text().strip()
        elif prompt_key not in config:
            raise FileNotFoundError(f"Missing prompt file for {prompt_key}: {prompt_path}")

    # Optional extra prompts (e.g. historical_prompt).
    for prompt_key, rel_path in prompt_files.items():
        if prompt_key in config:
            continue
        if not prompt_key.endswith("_prompt"):
            continue
        prompt_path = base_dir / rel_path
        if prompt_path.exists():
            config[prompt_key] = prompt_path.read_text().strip()
        else:
            raise FileNotFoundError(f"Missing prompt file for {prompt_key}: {prompt_path}")

    return config


# ─── Web Scraping ─────────────────────────────────────────────────────────────
def _safe_driver_quit(driver: webdriver.Chrome | None) -> None:
    if driver is None:
        return
    try:
        driver.quit()
    except Exception:
        pass


def build_driver() -> webdriver.Chrome:
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1400,2400")
    options.add_argument(f"--user-agent={SCRAPE_USER_AGENT}")
    if CHROMIUM_BINARY:
        options.binary_location = CHROMIUM_BINARY
    service = Service(CHROMEDRIVER_PATH) if CHROMEDRIVER_PATH else Service()
    driver = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(SCRAPE_TIMEOUT_SECONDS)
    return driver


def _get_with_retry(driver: webdriver.Chrome, url: str, retries: int = 2) -> bool:
    for attempt in range(retries + 1):
        try:
            driver.get(url)
            return True
        except (TimeoutException, WebDriverException) as exc:
            if attempt >= retries:
                logger.warning(f"Selenium failed for {url}: {exc}")
                return False
    return False


def _clean_scraped_text(text: str, max_chars: int = 1400) -> str:
    cleaned = " ".join((text or "").split())
    return cleaned[:max_chars]


def _query_terms(raw_query: str) -> str:
    query = re.sub(r"site:[^\s]+", "", raw_query, flags=re.IGNORECASE)
    query = re.sub(r"\b(OR|AND)\b", " ", query, flags=re.IGNORECASE)
    return " ".join(query.split())[:120]


def _selenium_item_quality_ok(snippet: str, title: str, link: str, source_key: str) -> bool:
    if not link:
        return False
    sk = source_key.lower()
    if sk == "forums":
        return len(title) >= 8 or len(snippet) >= 12
    if sk == "news":
        return (len(title) >= 5 and (len(snippet) >= 10 or len(title) >= 16)) or len(snippet) >= 40
    if sk == "twitter":
        return len(snippet) >= 10 or len(title) >= 8
    if sk == "reddit":
        return len(snippet) >= 12 or len(title) >= 10
    return not (len(snippet) < 25 and len(title) < 15)


def _finalize_selenium_items(raw: list[dict]) -> list[dict]:
    cap = SCRAPE_MAX_PAGES_PER_SOURCE * SCRAPE_MAX_RESULTS_PER_QUERY
    out: list[dict] = []
    for item in raw[:cap]:
        sk = (item.get("_sk") or "news").lower()
        snippet = _clean_scraped_text(item.get("snippet", ""))
        title = _clean_scraped_text(item.get("title", ""), max_chars=180)
        link = (item.get("link", "") or "").strip()
        if not _selenium_item_quality_ok(snippet, title, link, sk):
            continue
        out.append({
            "title": title or snippet[:120],
            "snippet": snippet or title,
            "link": link,
            "source": item.get("source", "News/Other"),
            "date": item.get("date", ""),
        })
    return out


def _extract_forums_from_driver(driver: webdriver.Chrome, query: str, max_results: int) -> list[dict]:
    items: list[dict] = []
    terms = _query_terms(query)
    if not terms:
        return []
    search_url = f"https://forums.redflagdeals.com/search.php?keywords={quote_plus(terms)}"
    if not _get_with_retry(driver, search_url):
        return []
    try:
        WebDriverWait(driver, SCRAPE_TIMEOUT_SECONDS).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "h3.searchresult-title, a.topictitle"))
        )
    except Exception:
        return []
    links = driver.find_elements(By.CSS_SELECTOR, "h3.searchresult-title a, a.topictitle")[:max_results]
    for el in links:
        title = _clean_scraped_text(el.text, max_chars=180)
        link = el.get_attribute("href") or ""
        if not title or not link:
            continue
        items.append({
            "title": title,
            "snippet": title,
            "link": link,
            "source": "RedFlagDeals",
            "date": "",
            "_sk": "forums",
        })
    return items


def _extract_x_from_driver(driver: webdriver.Chrome, query: str, max_results: int) -> list[dict]:
    items: list[dict] = []
    terms = _query_terms(query)
    if not terms:
        return []
    search_url = f"https://x.com/search?q={quote_plus(terms)}&src=typed_query&f=live"
    if not _get_with_retry(driver, search_url):
        return []
    try:
        WebDriverWait(driver, 8).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "article"))
        )
    except Exception:
        return []
    for tweet in driver.find_elements(By.CSS_SELECTOR, "article")[:max_results]:
        text_els = tweet.find_elements(By.CSS_SELECTOR, "[data-testid='tweetText']")
        link_els = tweet.find_elements(By.CSS_SELECTOR, "a[href*='/status/']")
        if not text_els or not link_els:
            continue
        body = _clean_scraped_text(text_els[0].text, max_chars=700)
        link = link_els[0].get_attribute("href") or ""
        if not body:
            continue
        items.append({
            "title": body[:120],
            "snippet": body,
            "link": link,
            "source": "X/Twitter",
            "date": "",
            "_sk": "twitter",
        })
    return items


def _unwrap_google_news_href(href: str) -> str:
    if not href:
        return ""
    if "url?q=" in href:
        try:
            q = parse_qs(urlparse(href).query).get("q", [""])[0]
            return q or href
        except Exception:
            return href
    if href.startswith("./"):
        return "https://news.google.com" + href[1:]
    return href


def _extract_news_from_driver(driver: webdriver.Chrome, query: str, max_results: int) -> list[dict]:
    items: list[dict] = []
    terms = _query_terms(query)
    if not terms:
        return []
    search_url = f"https://news.google.com/search?q={quote_plus(terms)}&hl=en-CA&gl=CA&ceid=CA%3Aen"
    if not _get_with_retry(driver, search_url):
        return []
    try:
        WebDriverWait(driver, SCRAPE_TIMEOUT_SECONDS).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "article"))
        )
    except Exception:
        return []
    for card in driver.find_elements(By.CSS_SELECTOR, "article"):
        if len(items) >= max_results:
            break
        href = ""
        title = ""
        anchor_els = card.find_elements(
            By.CSS_SELECTOR,
            "a[href*='./articles/'], a[href*='/articles/'], a[href*='url?q=']",
        )
        if not anchor_els:
            for a in card.find_elements(By.TAG_NAME, "a"):
                h = a.get_attribute("href") or ""
                if "articles" in h or "./articles" in h:
                    anchor_els = [a]
                    break
        if not anchor_els:
            continue
        title = _clean_scraped_text(anchor_els[0].text or "", max_chars=180)
        href = _unwrap_google_news_href(anchor_els[0].get_attribute("href") or "")
        snippet = _clean_scraped_text(card.text, max_chars=700)
        if not href:
            continue
        date_els = card.find_elements(By.CSS_SELECTOR, "time")
        raw_date = date_els[0].get_attribute("datetime") if date_els else ""
        items.append({
            "title": title or snippet[:140],
            "snippet": snippet or title,
            "link": href,
            "source": "News/Other",
            "date": _resolve_relative_date(raw_date),
            "_sk": "news",
        })
    return items


def _extract_reddit_selenium_from_driver(driver: webdriver.Chrome, query: str, max_results: int) -> list[dict]:
    items: list[dict] = []
    terms = _query_terms(query)
    if not terms:
        return []
    urls = (
        f"https://old.reddit.com/search/?q={quote_plus(terms)}&sort=new",
        f"https://old.reddit.com/search?q={quote_plus(terms)}&sort=new",
        f"https://www.reddit.com/search/?q={quote_plus(terms)}&sort=new&t=year",
    )
    for search_url in urls:
        if not _get_with_retry(driver, search_url):
            continue
        try:
            WebDriverWait(driver, min(SCRAPE_TIMEOUT_SECONDS, 22)).until(
                EC.presence_of_all_elements_located(
                    (By.CSS_SELECTOR, "div.search-result, div[data-testid='post-container'] a[href*='/comments/']")
                )
            )
        except Exception:
            continue
        cards = driver.find_elements(By.CSS_SELECTOR, "div.search-result")[:max_results]
        if not cards:
            links = driver.find_elements(
                By.CSS_SELECTOR, "a[href*='/r/'][href*='/comments/']"
            )[: max_results * 2]
            for a in links:
                title = _clean_scraped_text(a.text, max_chars=180)
                link = a.get_attribute("href") or ""
                if title and link and "/comments/" in link:
                    items.append({
                        "title": title,
                        "snippet": title,
                        "link": link,
                        "source": "Reddit",
                        "date": "",
                        "_sk": "reddit",
                    })
            if items:
                break
            continue
        for card in cards:
            title_els = card.find_elements(By.CSS_SELECTOR, "a.search-title, a[slot='title'], a.shreddit-post-title")
            if not title_els:
                continue
            title = _clean_scraped_text(title_els[0].text, max_chars=180)
            link = title_els[0].get_attribute("href") or ""
            snippet_els = card.find_elements(By.CSS_SELECTOR, "div.search-expando, div[slot='text-body']")
            snippet = _clean_scraped_text(snippet_els[0].text if snippet_els else "", max_chars=700)
            date_els = card.find_elements(By.CSS_SELECTOR, "time")
            raw_date = date_els[0].get_attribute("datetime") if date_els else ""
            items.append({
                "title": title,
                "snippet": snippet or title,
                "link": link,
                "source": "Reddit",
                "date": _resolve_relative_date(raw_date),
                "_sk": "reddit",
            })
        if items:
            break
    return items[:max_results]


def _scrape_et_bundle_sync(
    query: str,
    use_forums: bool,
    use_twitter: bool,
    use_news: bool,
    use_reddit_sel: bool,
    max_results: int,
) -> list[dict]:
    raw: list[dict] = []
    driver = None
    try:
        driver = build_driver()
        if use_forums:
            raw.extend(_extract_forums_from_driver(driver, query, max_results))
        if use_twitter and not _has_site_restriction(query):
            raw.extend(_extract_x_from_driver(driver, query, max_results))
        if use_news and not _has_site_restriction(query):
            raw.extend(_extract_news_from_driver(driver, query, max_results))
        if use_reddit_sel:
            raw.extend(_extract_reddit_selenium_from_driver(driver, query, max_results))
    finally:
        _safe_driver_quit(driver)
    return _finalize_selenium_items(raw)


def _scrape_comp_bundle_sync(
    query: str,
    use_forums: bool,
    use_twitter: bool,
    use_news: bool,
    use_reddit_sel: bool,
    max_results: int,
) -> list[dict]:
    raw: list[dict] = []
    driver = None
    try:
        driver = build_driver()
        if use_news:
            raw.extend(_extract_news_from_driver(driver, query, max_results))
        if use_twitter and not _has_site_restriction(query):
            raw.extend(_extract_x_from_driver(driver, query, max_results))
        if use_forums and ("forum" in query.lower() or "redflagdeals" in query.lower()):
            raw.extend(_extract_forums_from_driver(driver, query, max_results))
        if use_reddit_sel:
            raw.extend(_extract_reddit_selenium_from_driver(driver, query, max_results))
    finally:
        _safe_driver_quit(driver)
    return _finalize_selenium_items(raw)


def _scrape_reddit_source(query: str, max_results: int = 5) -> list[dict]:
    driver = None
    try:
        driver = build_driver()
        raw = _extract_reddit_selenium_from_driver(driver, query, max_results)
        return _finalize_selenium_items(raw)
    finally:
        _safe_driver_quit(driver)


def _scrape_forums_source(query: str, max_results: int = 5) -> list[dict]:
    driver = None
    try:
        driver = build_driver()
        raw = _extract_forums_from_driver(driver, query, max_results)
        return _finalize_selenium_items(raw)
    finally:
        _safe_driver_quit(driver)


def _scrape_x_source(query: str, max_results: int = 5) -> list[dict]:
    driver = None
    try:
        driver = build_driver()
        raw = _extract_x_from_driver(driver, query, max_results)
        return _finalize_selenium_items(raw)
    finally:
        _safe_driver_quit(driver)


def _scrape_news_source(query: str, max_results: int = 5) -> list[dict]:
    driver = None
    try:
        driver = build_driver()
        raw = _extract_news_from_driver(driver, query, max_results)
        return _finalize_selenium_items(raw)
    finally:
        _safe_driver_quit(driver)


async def selenium_source_search(
    query: str,
    source: str,
    max_results: int = 5,
) -> list[dict]:
    source_key = source.lower()
    if source_key == "reddit":
        fn = _scrape_reddit_source
    elif source_key == "forums":
        fn = _scrape_forums_source
    elif source_key == "twitter":
        fn = _scrape_x_source
    elif source_key == "news":
        fn = _scrape_news_source
    else:
        return []
    async with _selenium_browser_sem:
        results = await asyncio.to_thread(fn, query, min(max_results, SCRAPE_MAX_RESULTS_PER_QUERY))
    logger.info(f"Selenium source={source} query='{query[:40]}' -> {len(results)} mentions")
    return results


async def selenium_et_bundle(query: str, use_forums: bool, use_twitter: bool, use_news: bool, use_reddit_sel: bool) -> list[dict]:
    mr = SCRAPE_MAX_RESULTS_PER_QUERY
    async with _selenium_browser_sem:
        return await asyncio.to_thread(
            _scrape_et_bundle_sync, query, use_forums, use_twitter, use_news, use_reddit_sel, mr
        )


async def selenium_comp_bundle(query: str, use_forums: bool, use_twitter: bool, use_news: bool, use_reddit_sel: bool) -> list[dict]:
    mr = SCRAPE_MAX_RESULTS_PER_QUERY
    async with _selenium_browser_sem:
        return await asyncio.to_thread(
            _scrape_comp_bundle_sync, query, use_forums, use_twitter, use_news, use_reddit_sel, mr
        )


def lookback_hours_to_tbs(lookback_hours: int) -> str:
    # Google-style time filters used by query config.
    if lookback_hours <= 24:
        return "qdr:d"
    if lookback_hours <= 24 * 7:
        return "qdr:w"
    return "qdr:m"


def normalize_tbs(tbs: str) -> str:
    """Normalize tbs value. Empty string means 'all time' (no time filter)."""
    if tbs in ("", "all"):
        return ""
    supported = {"qdr:d", "qdr:w", "qdr:m", "qdr:y"}
    return tbs if tbs in supported else "qdr:m"


def _has_site_restriction(query: str) -> bool:
    return "site:" in query.lower()


_search_errors: list[str] = []


def _classify_channel_and_source(link: str) -> tuple[str, str]:
    url = (link or "").lower()
    if "reddit.com" in url:
        return "people", "Reddit"
    if "x.com" in url or "twitter.com" in url:
        return "people", "X/Twitter"
    if "redflagdeals.com" in url:
        return "people", "RedFlagDeals"
    if "forum" in url or "community" in url:
        return "people", "Forum"
    return "press", "News/Other"


def _source_quality_tier(link: str, channel: str) -> str:
    url = (link or "").lower()
    if channel == "people":
        return "tier1_user_generated"
    if any(d in url for d in ["reddit.com", "x.com", "twitter.com", "redflagdeals.com"]):
        return "tier1_user_generated"
    if any(
        d in url
        for d in [
            "reuters.com",
            "bloomberg.com",
            "cbc.ca",
            "theglobeandmail.com",
            "financialpost.com",
        ]
    ):
        return "tier2_reported"
    return "tier3_commentary_or_unknown"


def _detect_brands(text: str) -> str:
    content = (text or "").lower()
    brand_order = [
        ("interac", "Interac"),
        ("wise", "Wise"),
        ("paypal", "PayPal"),
        ("apple pay", "ApplePay"),
        ("google pay", "GooglePay"),
        ("samsung pay", "SamsungPay"),
        ("venmo", "Venmo"),
        ("cash app", "CashApp"),
    ]
    found = [name for token, name in brand_order if token in content]
    return ", ".join(found) if found else "Unknown"


def _detect_use_case(text: str) -> str:
    content = (text or "").lower()
    if any(x in content for x in ["cross-border", "outside canada", "international", "remittance"]):
        return "cross_border_transfer"
    if any(x in content for x in ["fraud", "scam", "security", "hold", "risk"]):
        return "fraud_assurance"
    if any(x in content for x in ["wallet", "apple pay", "google pay", "checkout", "tap"]):
        return "wallet_or_checkout"
    if any(x in content for x in ["business", "payroll", "merchant"]):
        return "business_payment"
    if any(x in content for x in ["delay", "slow", "instant", "speed", "pending", "transfer"]):
        return "domestic_transfer_speed"
    return "general_payments"


_USE_CASE_LABELS = {
    "cross_border_transfer": "Cross-Border Transfers",
    "fraud_assurance": "Fraud Confidence",
    "wallet_or_checkout": "Wallet And Checkout",
    "business_payment": "Business Payments",
    "domestic_transfer_speed": "Domestic Transfer Speed",
    "general_payments": "General Payments",
}


def _use_case_label(use_case: str) -> str:
    return _USE_CASE_LABELS.get(use_case, use_case.replace("_", " ").title())


def _corroboration_label(unique_domains: int) -> str:
    if unique_domains >= 3:
        return "strong"
    if unique_domains == 2:
        return "moderate"
    return "early"


_CLUSTER_STOPWORDS = {
    "the",
    "and",
    "for",
    "with",
    "that",
    "this",
    "from",
    "into",
    "their",
    "about",
    "your",
    "have",
    "has",
    "had",
    "are",
    "was",
    "were",
    "will",
    "would",
    "should",
    "could",
    "can",
    "but",
    "not",
    "you",
    "they",
    "its",
    "it's",
    "canada",
    "payment",
    "payments",
}


def _normalize_date_value(raw_date: str) -> str:
    value = (raw_date or "").strip()
    if not value:
        return "unknown"
    iso_match = re.search(r"\b(\d{4}-\d{2}-\d{2})\b", value)
    if iso_match:
        return iso_match.group(1)
    return value[:24]


def _extract_date_from_url(url: str) -> str:
    """Layer 1: extract publication date from common URL path patterns.

    Covers most news sites that embed /YYYY/MM/DD/ in their URL structure
    (CBC, Globe and Mail, TechCrunch, Reuters, etc.).
    Returns ISO "YYYY-MM-DD" string or empty string if no date found.
    """
    # /YYYY/MM/DD/ path pattern
    m = re.search(r"/(\d{4})/(\d{1,2})/(\d{1,2})(?:/|$)", url)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 2000 <= y <= 2030 and 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{y:04d}-{mo:02d}-{d:02d}"
    # YYYY-MM-DD or YYYY_MM_DD in URL query string or path segment
    m = re.search(r"(?<!\d)(\d{4})[_\-](\d{2})[_\-](\d{2})(?!\d)", url)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 2000 <= y <= 2030 and 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{y:04d}-{mo:02d}-{d:02d}"
    return ""


def _extract_date_from_snippet(text: str) -> str:
    """Layer 2: extract publication date from snippet/body text.

    Handles ISO dates and common English date phrases already present in
    the search snippet (e.g. "Published April 10, 2025" or "2025-03-15").
    Returns ISO "YYYY-MM-DD" string or empty string if nothing matched.
    """
    if not text:
        return ""
    # ISO date in text
    m = re.search(r"\b(\d{4}-\d{2}-\d{2})\b", text)
    if m:
        return m.group(1)
    # "Month DD, YYYY" or "Month DD YYYY"
    _months = (
        "January|February|March|April|May|June|July|August|"
        "September|October|November|December|"
        "Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Oct|Nov|Dec"
    )
    m = re.search(
        rf"\b({_months})\s+(\d{{1,2}}),?\s+(\d{{4}})\b", text, re.IGNORECASE
    )
    if m:
        for fmt in ("%B %d %Y", "%b %d %Y"):
            try:
                dt = datetime.strptime(
                    f"{m.group(1)} {m.group(2)} {m.group(3)}", fmt
                )
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue
    return ""


def _extract_keywords(text: str, max_keywords: int = 6) -> list[str]:
    tokens = re.findall(r"[a-z0-9]+", (text or "").lower())
    keywords: list[str] = []
    for token in tokens:
        if len(token) < 3 or token in _CLUSTER_STOPWORDS:
            continue
        if token not in keywords:
            keywords.append(token)
        if len(keywords) >= max_keywords:
            break
    return keywords


def _cluster_key_from_components(brands: str, use_case: str, keywords: list[str]) -> str:
    keyword_key = ",".join(sorted(keywords[:4])) if keywords else "nokeywords"
    return f"{brands}|{use_case}|{keyword_key}"


def _token_overlap_ratio(tokens_a: set[str], tokens_b: set[str]) -> float:
    if not tokens_a or not tokens_b:
        return 0.0
    overlap = len(tokens_a.intersection(tokens_b))
    return overlap / float(min(len(tokens_a), len(tokens_b)))


def _cluster_mentions(mentions: list[dict]) -> list[dict]:
    clusters: list[dict] = []
    for mention in mentions:
        mention_tokens = set(mention.get("keywords", []))
        chosen_cluster = None

        for cluster in clusters:
            rep = cluster["representative"]
            rep_tokens = set(rep.get("keywords", []))
            overlap = _token_overlap_ratio(mention_tokens, rep_tokens)

            same_brand = mention.get("brands", "Unknown") == rep.get("brands", "Unknown")
            same_use_case = mention.get("use_case", "general_payments") == rep.get("use_case", "general_payments")

            # Anti-fragmentation guard:
            # - strong overlap always merges
            # - fallback merge for same brand + use-case with moderate overlap
            if overlap >= 0.6 or (same_brand and same_use_case and overlap >= 0.35):
                chosen_cluster = cluster
                break

        if chosen_cluster is None:
            chosen_cluster = {
                "cluster_key": mention["cluster_key"],
                "mentions": [],
                "representative": mention,
            }
            clusters.append(chosen_cluster)

        chosen_cluster["mentions"].append(mention)
        if len(mention.get("snippet", "")) > len(chosen_cluster["representative"].get("snippet", "")):
            chosen_cluster["representative"] = mention

    summarized = []
    for idx, cluster in enumerate(clusters, 1):
        cluster_mentions = cluster["mentions"]
        rep = cluster["representative"]
        domains = sorted({urlparse(m.get("url", "")).netloc.replace("www.", "") for m in cluster_mentions if m.get("url")})
        timeframes = sorted({m.get("timeframe", "unknown") for m in cluster_mentions})
        known_dates = sorted(
            {
                d for d in (m.get("date", "unknown") for m in cluster_mentions)
                if d and d != "unknown"
            }
        )
        dated_count = sum(1 for m in cluster_mentions if m.get("date", "unknown") != "unknown")
        if known_dates:
            date_span = known_dates[0] if len(known_dates) == 1 else f"{known_dates[0]} to {known_dates[-1]}"
        else:
            date_span = "unknown"
        summarized.append(
            {
                "story_id": f"S{idx}",
                "archetype_hint": _use_case_label(rep.get("use_case", "general_payments")),
                "brands": rep.get("brands", "Unknown"),
                "article_count": len(cluster_mentions),
                "unique_domains": len(domains),
                "corroboration": _corroboration_label(len(domains)),
                "timeframes_present": ", ".join(timeframes) if timeframes else "unknown",
                "dated_count": dated_count,
                "date_span": date_span,
                "sample_urls": [m.get("url", "") for m in cluster_mentions[:3] if m.get("url")],
                "sample_snippet": rep.get("snippet", "")[:220],
            }
        )

    summarized.sort(key=lambda c: (c["article_count"], c["unique_domains"], c["dated_count"]), reverse=True)
    return summarized


def _extract_platform_context(link: str) -> dict[str, str]:
    """Extract persona-relevant metadata from mention URLs."""
    url = (link or "").lower()
    ctx = {
        "subreddit": "",
        "forum_section": "",
        "platform_demo_hint": "",
    }

    subreddit_match = re.search(r"reddit\.com/r/([a-z0-9_]+)", url)
    if subreddit_match:
        subreddit = subreddit_match.group(1)
        subreddit_hints = {
            "personalfinancecanada": "personal finance consumer, likely 25-45",
            "canadianinvestor": "investor, likely 30-55",
            "canada": "general Canadian public",
            "ontario": "Ontario resident",
        }
        ctx["subreddit"] = subreddit
        ctx["platform_demo_hint"] = subreddit_hints.get(
            subreddit,
            "Reddit community user, likely detail-oriented and price-sensitive",
        )
        return ctx

    if "redflagdeals.com" in url:
        ctx["forum_section"] = "RedFlagDeals"
        ctx["platform_demo_hint"] = "deal-seeking consumer, budget-conscious, likely 25-45"
        return ctx

    if "x.com" in url or "twitter.com" in url:
        ctx["forum_section"] = "X/Twitter"
        ctx["platform_demo_hint"] = "social media user, skews 20-40, more reactive"
        return ctx

    if "forum" in url or "community" in url:
        ctx["forum_section"] = "Forum/Community"
        ctx["platform_demo_hint"] = "community forum user, likely troubleshooting-focused"

    return ctx


def _tbs_to_timelimit(tbs: str) -> str | None:
    mapping = {"qdr:d": "d", "qdr:w": "w", "qdr:m": "m", "qdr:y": "y"}
    return mapping.get(tbs) if tbs else None


def _resolve_relative_date(date_str: str, *, tbs: str = "") -> str:
    """Normalize DDG date strings to "Month DD, YYYY" format (UTC dates).

    DDG text() returns relative strings like '3 weeks ago', '2 months ago'.
    DDG news() returns ISO timestamps or pre-formatted strings.
    All dates are kept in UTC — the native timezone of Reddit's created_utc
    and DDG's timestamps — so the displayed date matches when the content
    was actually published.
    Returns empty string if no date is available.
    """
    if not date_str:
        return date_str
    # ISO dates: "2025-04-10", "2025-04-10T14:22:00+00:00", "2025-04-10 14:22:00"
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", date_str)
    if m:
        try:
            dt = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)), tzinfo=timezone.utc)
            return dt.strftime("%B %d, %Y")
        except ValueError:
            pass
    # RFC 2822: "Mon, 14 Apr 2026 01:30:00 +0000"
    try:
        from email.utils import parsedate_to_datetime as _pdt
        dt = _pdt(date_str).astimezone(timezone.utc)
        return dt.strftime("%B %d, %Y")
    except Exception:
        pass
    # Named-month: "April 14, 2026" / "Apr 14, 2026" / "14 Apr 2026"
    for fmt in ("%B %d, %Y", "%b %d, %Y", "%d %B %Y", "%d %b %Y"):
        try:
            return datetime.strptime(date_str.strip(), fmt).strftime("%B %d, %Y")
        except ValueError:
            pass
    # Parse relative strings using UTC now so the computed date matches the source
    now = datetime.now(timezone.utc)
    m = re.match(r"(\d+)\s+(day|week|month|year)s?\s+ago", date_str.strip(), re.IGNORECASE)
    if m:
        n = int(m.group(1))
        unit = m.group(2).lower()
        if unit == "day":
            dt = now - timedelta(days=n)
        elif unit == "week":
            dt = now - timedelta(weeks=n)
        elif unit == "month":
            dt = now - timedelta(days=n * 30)
        else:  # year
            dt = now - timedelta(days=n * 365)
        return dt.strftime("%B %d, %Y")
    if re.match(r"\d+\s+(hour|minute)s?\s+ago", date_str.strip(), re.IGNORECASE):
        return now.strftime("%B %d, %Y")
    return date_str  # unknown format — keep as-is


async def _fetch_meta_date(url: str, client: httpx.AsyncClient) -> str:
    """Layer 3: scrape HTML <head> for Open Graph / JSON-LD / <time> publication date.

    Returns "Month DD, YYYY" string or empty string on failure.
    Only reads the first 6 KB of the response — enough to cover <head>.
    """
    try:
        resp = await client.get(
            url,
            follow_redirects=True,
            timeout=4.0,
            headers={"User-Agent": "Mozilla/5.0 (compatible; InteracIntelBot/1.0)"},
        )
        if resp.status_code != 200:
            return ""
        head_html = resp.text[:6000]
    except Exception:
        return ""

    patterns = [
        # Open Graph article:published_time (both attribute orderings)
        r'<meta[^>]+property=["\']article:published_time["\'][^>]+content=["\']([\d\-T:+Z]+)["\']',
        r'<meta[^>]+content=["\']([\d\-T:+Z]+)["\'][^>]+property=["\']article:published_time["\']',
        # Generic pubdate / DC.date meta tags
        r'<meta[^>]+name=["\'](?:pubdate|DC\.date)["\'][^>]+content=["\']([\d\-T:+Z]+)["\']',
        # JSON-LD datePublished
        r'"datePublished"\s*:\s*"([\d\-T:+Z]+)"',
        # HTML5 <time datetime="...">
        r'<time[^>]+datetime=["\']([\d\-T:+Z]+)["\']',
    ]
    for pat in patterns:
        m = re.search(pat, head_html, re.IGNORECASE)
        if m:
            resolved = _resolve_relative_date(m.group(1))
            if resolved and not re.match(r"^\d{4}-\d{2}-\d{2}T", resolved):
                # _resolve_relative_date already formatted it as "Month DD, YYYY"
                return resolved
            # ISO timestamp that _resolve_relative_date parsed cleanly
            if resolved:
                return resolved
    return ""


async def _enrich_dates_from_meta(
    mentions: list[dict], max_fetches: int = 20
) -> None:
    """Layer 3 batch enrichment: fetch HTML meta dates for undated mentions in-place.

    Skips mentions that already have a date. Caps HTTP requests at max_fetches
    to bound latency (each request has a 4 s timeout, all run concurrently).
    """
    undated = [
        m for m in mentions
        if not m.get("date") and m.get("link")
    ]
    targets = undated[:max_fetches]
    if not targets:
        return

    async with httpx.AsyncClient() as client:
        results = await asyncio.gather(
            *[_fetch_meta_date(m["link"], client) for m in targets],
            return_exceptions=True,
        )

    gained = 0
    for mention, result in zip(targets, results):
        if isinstance(result, str) and result:
            mention["date"] = result
            gained += 1

    logger.info(f"[date-enrich-L3] fetched {len(targets)} undated URLs, gained {gained} dates")


async def web_search(
    query: str,
    search_type: str = "search",
    max_results: int = 5,
    tbs: str = "qdr:w",
) -> list[dict]:
    timelimit = _tbs_to_timelimit(tbs)

    def _run_search() -> list[dict]:
        with DDGS() as ddgs:
            if search_type == "news":
                raw = list(ddgs.news(query, max_results=max_results, timelimit=timelimit))
            else:
                raw = list(ddgs.text(query, max_results=max_results, timelimit=timelimit))
        normalized = []
        for item in raw:
            link = item.get("href", "") or item.get("url", "")
            # text() uses "published" ("2 weeks ago"); news() uses "date" (ISO).
            raw_date = item.get("date", "") or item.get("published", "")
            # Layer 1: try URL path pattern when API field is empty
            if not raw_date:
                raw_date = _extract_date_from_url(link)
            # Layer 2: try snippet text when URL pattern also fails
            if not raw_date:
                raw_date = _extract_date_from_snippet(
                    item.get("body", "") or item.get("snippet", "")
                )
            normalized.append({
                "title": item.get("title", ""),
                "snippet": item.get("body", "") or item.get("snippet", ""),
                "link": link,
                "source": item.get("source", search_type),
                # _resolve_relative_date converts any date string to "Month DD, YYYY".
                # Pass tbs so empty dates get an approximate fallback from the search window.
                "date": _resolve_relative_date(raw_date, tbs=tbs),
            })
        return normalized

    try:
        results = await asyncio.to_thread(_run_search)
    except Exception as e:
        err = f"DDG exception for [{search_type}] '{query[:40]}': {type(e).__name__}: {e}"
        logger.error(err)
        _search_errors.append(err)
        return []

    logger.info(f"DDG [{search_type}] '{query}' tbs={tbs!r} -> {len(results)} results")
    return results


async def search_twitter(query: str, max_results: int = 5, tbs: str = "qdr:w") -> list[dict]:
    """Search X/Twitter via DDG. Skips if query already has a site: restriction."""
    if _has_site_restriction(query):
        return []
    base_results = await web_search(
        f"{query} site:x.com OR site:twitter.com",
        "search",
        max_results=max_results,
        tbs=tbs,
    )
    for r in base_results:
        r["source"] = "X/Twitter"
    return base_results


_REDDIT_HEADERS = {
    "User-Agent": "python:interac.intelligence.bot:v1.0 (by /u/interac_intel_bot)"
}


def _parse_reddit_post(post: dict, cutoff_ts: float) -> dict | None:
    """Parse a Reddit post dict into our standard format. Returns None if filtered out."""
    created = post.get("created_utc", 0)
    if not created or created < cutoff_ts:
        return None
    title = (post.get("title", "") or "").strip()
    selftext = (post.get("selftext", "") or "").strip()
    if selftext in ("[deleted]", "[removed]"):
        selftext = ""
    permalink = post.get("permalink", "")
    subreddit_name = post.get("subreddit_display_name", post.get("subreddit", ""))
    post_dt = datetime.fromtimestamp(created, tz=timezone.utc)
    return {
        "title": title,
        "snippet": selftext[:600] if selftext else title,
        "link": f"https://www.reddit.com{permalink}",
        "source": f"Reddit/r/{subreddit_name}" if subreddit_name else "Reddit",
        "date": post_dt.strftime("%B %d, %Y"),
        "score": post.get("score", 0),
        "permalink": permalink,
    }


async def _reddit_get(url: str, params: dict) -> dict | None:
    """GET a Reddit JSON endpoint using httpx (urllib gets 403'd by Reddit's TLS check)."""
    try:
        async with httpx.AsyncClient(timeout=15, headers=_REDDIT_HEADERS) as client:
            r = await client.get(url, params=params)
        if r.status_code != 200:
            logger.warning(f"Reddit API {r.status_code} for {url}")
            return None
        return r.json()
    except Exception as e:
        logger.warning(f"Reddit API failed for {url}: {type(e).__name__}: {e}")
        return None


async def fetch_reddit_comments(permalink: str, max_comments: int = 6) -> list[str]:
    """Fetch top comments from a post by appending .json to its permalink URL.

    This is the 'add .json to any Reddit URL' trick — gives full post + comments.
    Only returns comments with score >= 2.
    """
    url = f"https://www.reddit.com{permalink}.json"
    try:
        async with httpx.AsyncClient(timeout=12, headers=_REDDIT_HEADERS) as client:
            r = await client.get(url, params={"limit": max_comments, "sort": "top"})
        if r.status_code != 200:
            return []
        data = r.json()
        # data[0] = post listing, data[1] = comment listing
        comments = data[1]["data"]["children"]
        result = []
        for c in comments[:max_comments]:
            d = c.get("data", {})
            body = (d.get("body", "") or "").strip()
            cscore = d.get("score", 0)
            if body and body not in ("[deleted]", "[removed]") and cscore >= 2:
                result.append(body[:350])
        return result
    except Exception as e:
        logger.debug(f"Comment fetch failed for {permalink}: {e}")
        return []


async def search_reddit_posts(
    query: str,
    subreddit: str = "",
    max_results: int = 15,
    days_back: int = 60,
    min_score: int = 1,
    enrich_comments: bool = False,
) -> list[dict]:
    """Search Reddit posts via the public JSON API using httpx.

    - Exact UTC timestamps → real dates
    - Full post selftext → better quote material
    - Optional comment enrichment: for posts with score >= 5, fetches top comment
      (the .json trick) and appends it to the snippet so the LLM has community
      reactions, not just the OP's question
    """
    if subreddit:
        url = f"https://www.reddit.com/r/{subreddit}/search.json"
        params: dict = {
            "q": query, "sort": "new", "t": "year",
            "limit": min(max_results * 2, 25), "restrict_sr": "1", "type": "link",
        }
    else:
        url = "https://www.reddit.com/search.json"
        params = {
            "q": query, "sort": "new", "t": "year",
            "limit": min(max_results * 2, 25), "type": "link",
        }

    cutoff_ts = (datetime.now(timezone.utc) - timedelta(days=days_back)).timestamp()
    data = await _reddit_get(url, params)
    if not data:
        return []

    posts = []
    for child in data.get("data", {}).get("children", []):
        parsed = _parse_reddit_post(child.get("data", {}), cutoff_ts)
        if parsed and parsed["score"] >= min_score:
            posts.append(parsed)

    posts.sort(key=lambda x: x["score"], reverse=True)
    posts = posts[:max_results]

    # Enrich top posts with their best comment via the .json permalink trick
    if enrich_comments:
        enrichment_tasks = []
        for p in posts:
            if p["score"] >= 5 and p.get("permalink"):
                enrichment_tasks.append((p, fetch_reddit_comments(p["permalink"], max_comments=4)))
            else:
                enrichment_tasks.append((p, None))

        for post, coro in enrichment_tasks:
            if coro is None:
                continue
            comments = await coro
            if comments:
                # Prepend the best comment to give the LLM community reaction context
                post["snippet"] = post["snippet"] + "\n\nTop community reply: " + comments[0]

    return posts


async def browse_subreddit_new(
    subreddit: str,
    keyword: str,
    days_back: int = 30,
    limit: int = 100,
) -> list[dict]:
    """Browse a subreddit's /new feed and filter by keyword — no search API bias.

    Useful when you want all recent posts mentioning a term, not just what
    Reddit's search algorithm returns.
    """
    url = f"https://www.reddit.com/r/{subreddit}/new.json"
    cutoff_ts = (datetime.now(timezone.utc) - timedelta(days=days_back)).timestamp()
    data = await _reddit_get(url, {"limit": limit})
    if not data:
        return []

    kw = keyword.lower()
    posts = []
    for child in data.get("data", {}).get("children", []):
        post_data = child.get("data", {})
        title = (post_data.get("title", "") or "").lower()
        selftext = (post_data.get("selftext", "") or "").lower()
        if kw not in title and kw not in selftext:
            continue
        parsed = _parse_reddit_post(post_data, cutoff_ts)
        if parsed:
            posts.append(parsed)

    posts.sort(key=lambda x: x["score"], reverse=True)
    return posts


async def _search_diagnostic() -> str:
    """Single test query to diagnose DDG search health."""
    try:
        results = await web_search("Interac e-Transfer", "search", max_results=3, tbs="")
    except Exception as e:
        return f"FAIL: {type(e).__name__}: {e}"
    if not results:
        return "FAIL: no search results returned"
    return f"OK: {len(results)} results, first='{results[0].get('title', 'n/a')[:80]}'"


async def fetch_biweekly_mentions() -> str:
    """Fetch mentions for the universal biweekly scan.

    Strategy:
    - Reddit API (primary): exact dates, full post content, quality score filter.
      Used for both e-Transfer pain posts and competitor community reactions.
    - DDG (supplement): RFD, X/Twitter, news for things Reddit API can't cover.
    Results split into labelled sections so the LLM assigns them correctly.
    """
    config = load_prompts()
    etransfer_ddg_queries = config.get("etransfer_queries", config.get("biweekly_queries", []))
    competitor_ddg_queries = config.get("competitor_queries", [])
    source_toggles = config.get("sources", {})
    use_reddit = bool(source_toggles.get("reddit", True))
    use_forums = bool(source_toggles.get("forums", True))
    use_twitter = bool(source_toggles.get("twitter", True))
    use_news = bool(source_toggles.get("news", True))

    seen_links: set[str] = set()
    etransfer_social: list[dict] = []
    etransfer_press: list[dict] = []
    competitor_mentions: list[dict] = []

    sem = asyncio.Semaphore(3)  # Max 3 concurrent Reddit API calls

    async def _search(q: str, sub: str, days: int, score: int = 1, comments: bool = False) -> list[dict]:
        async with sem:
            return await search_reddit_posts(
                q, subreddit=sub, max_results=12, days_back=days,
                min_score=score, enrich_comments=comments,
            )

    async def _browse(sub: str, kw: str, days: int) -> list[dict]:
        async with sem:
            return await browse_subreddit_new(sub, kw, days_back=days, limit=100)

    # ── 1. Reddit — e-Transfer community ──
    # Search API for targeted queries + browse /new feed for anything mentioning
    # e-transfer that search might miss (no search-ranking bias on the browse).
    reddit_et_searches = [
        ("e-transfer", "personalfinancecanada"),
        ("interac e-transfer", "personalfinancecanada"),
        ("e-transfer fraud OR scam", ""),
        ("e-transfer problem OR issue OR delay OR complaint", ""),
        ("e-transfer limit OR hold OR declined OR pending", ""),
        ("interac e-transfer", "canada"),
        ("e-transfer", "banking"),
    ]
    et_browse = [
        ("personalfinancecanada", "e-transfer", 45),
        ("personalfinancecanada", "interac", 45),
        ("canada", "e-transfer", 30),
    ]

    reddit_et_tasks_run = 0
    reddit_et_exceptions = 0
    reddit_et_post_rows = 0
    after_reddit_et_only = 0
    after_reddit_comp_only = 0
    if use_reddit:
        et_search_tasks = [_search(q, sub, 60) for q, sub in reddit_et_searches]
        et_browse_tasks = [_browse(sub, kw, days) for sub, kw, days in et_browse]
        et_batches = await asyncio.gather(*et_search_tasks, *et_browse_tasks, return_exceptions=True)
        reddit_et_tasks_run = len(et_batches)
        for batch in et_batches:
            if isinstance(batch, Exception):
                reddit_et_exceptions += 1
                continue
            if not batch:
                continue
            reddit_et_post_rows += len(batch)
            for r in batch:
                link = r.get("link", "")
                if link and link not in seen_links:
                    seen_links.add(link)
                    etransfer_social.append(r)
        after_reddit_et_only = len(etransfer_social)

    # ── 2. Reddit — competitor community reactions (with comment enrichment) ──
    # enrich_comments=True: for posts score>=5, fetches top comment via .json trick
    # so the LLM gets community reaction, not just the OP's question.
    reddit_comp_searches = [
        ("wise money transfer canada", "personalfinancecanada"),
        ("paypal canada send money", "personalfinancecanada"),
        ("wealthsimple cash transfer", "personalfinancecanada"),
        ("koho card banking", "personalfinancecanada"),
        ("best way send money canada interac alternative", "personalfinancecanada"),
        ("apple pay google pay canada", "personalfinancecanada"),
        ("revolut canada", "personalfinancecanada"),
        ("neo financial banking canada", "personalfinancecanada"),
        ("venmo paypal zelle canada", ""),
        ("digital wallet canada", "personalfinancecanada"),
    ]
    comp_browse = [
        ("personalfinancecanada", "wise", 90),
        ("personalfinancecanada", "paypal", 90),
        ("personalfinancecanada", "wealthsimple", 60),
    ]

    reddit_comp_tasks_run = 0
    reddit_comp_exceptions = 0
    reddit_comp_post_rows = 0
    if use_reddit:
        comp_search_tasks = [_search(q, sub, 180, score=1, comments=True) for q, sub in reddit_comp_searches]
        comp_browse_tasks = [_browse(sub, kw, days) for sub, kw, days in comp_browse]
        comp_batches = await asyncio.gather(*comp_search_tasks, *comp_browse_tasks, return_exceptions=True)
        reddit_comp_tasks_run = len(comp_batches)
        for batch in comp_batches:
            if isinstance(batch, Exception):
                reddit_comp_exceptions += 1
                continue
            if not batch:
                continue
            reddit_comp_post_rows += len(batch)
            for r in batch:
                link = r.get("link", "")
                if link and link not in seen_links:
                    seen_links.add(link)
                    competitor_mentions.append(r)
        after_reddit_comp_only = len(competitor_mentions)

    # ── 3. Selenium (one browser per query round) — e-Transfer supplement ──
    selenium_et_query_count = 0
    selenium_et_rows = 0
    for query in etransfer_ddg_queries:
        selenium_et_query_count += 1
        rows = await selenium_et_bundle(
            query,
            use_forums,
            use_twitter,
            use_news,
            bool(use_reddit and _has_site_restriction(query)),
        )
        selenium_et_rows += len(rows)
        for r in rows:
            link = r.get("link", "")
            if not link or link in seen_links:
                continue
            seen_links.add(link)
            channel, source = _classify_channel_and_source(link)
            r["channel"] = channel
            r["source"] = source
            if channel == "people":
                etransfer_social.append(r)
            else:
                etransfer_press.append(r)

    # ── 4. Selenium — competitor press/news + X ──
    selenium_comp_query_count = 0
    selenium_comp_rows = 0
    for query in competitor_ddg_queries:
        selenium_comp_query_count += 1
        rows = await selenium_comp_bundle(
            query,
            use_forums,
            use_twitter,
            use_news,
            bool(use_reddit and _has_site_restriction(query)),
        )
        selenium_comp_rows += len(rows)
        for r in rows:
            link = r.get("link", "")
            if not link or link in seen_links:
                continue
            seen_links.add(link)
            _, source = _classify_channel_and_source(link)
            r["source"] = source
            competitor_mentions.append(r)

    total_before_ddg = len(etransfer_social) + len(etransfer_press) + len(competitor_mentions)
    ddg_fallback_used = False
    ddg_rows_added = 0
    if FETCH_FALLBACK_DDG and total_before_ddg == 0:
        ddg_fallback_used = True
        for query in etransfer_ddg_queries:
            for r in await web_search(query, "search", 5, tbs="qdr:m"):
                link = r.get("link", "")
                if not link or link in seen_links:
                    continue
                seen_links.add(link)
                ddg_rows_added += 1
                channel, source = _classify_channel_and_source(link)
                r["channel"] = channel
                r["source"] = source
                if channel == "people":
                    etransfer_social.append(r)
                else:
                    etransfer_press.append(r)
            if not _has_site_restriction(query):
                for r in await web_search(query, "news", 5, tbs="qdr:m"):
                    link = r.get("link", "")
                    if not link or link in seen_links:
                        continue
                    seen_links.add(link)
                    ddg_rows_added += 1
                    channel, source = _classify_channel_and_source(link)
                    r["channel"] = channel
                    r["source"] = source
                    if channel == "people":
                        etransfer_social.append(r)
                    else:
                        etransfer_press.append(r)
                for r in await search_twitter(query, 5, tbs="qdr:m"):
                    link = r.get("link", "")
                    if not link or link in seen_links:
                        continue
                    seen_links.add(link)
                    ddg_rows_added += 1
                    r["channel"] = "people"
                    r["source"] = "X/Twitter"
                    etransfer_social.append(r)
        for query in competitor_ddg_queries:
            for search_type, tbs in [("search", "qdr:y"), ("news", "qdr:y")]:
                for r in await web_search(query, search_type, 5, tbs=tbs):
                    link = r.get("link", "")
                    if not link or link in seen_links:
                        continue
                    seen_links.add(link)
                    ddg_rows_added += 1
                    _, source = _classify_channel_and_source(link)
                    r["source"] = source
                    competitor_mentions.append(r)
            if not _has_site_restriction(query):
                for r in await search_twitter(query, 5, tbs="qdr:y"):
                    link = r.get("link", "")
                    if not link or link in seen_links:
                        continue
                    seen_links.add(link)
                    ddg_rows_added += 1
                    r["source"] = "X/Twitter"
                    competitor_mentions.append(r)

    # Layer 3: async HTML meta scraping for any DDG results still missing a date.
    # Reddit mentions already have exact dates; this targets DDG text() results.
    # Run across all three buckets — Reddit items are skipped automatically
    # because they already have a non-empty date field.
    all_mentions = etransfer_social + etransfer_press + competitor_mentions
    await _enrich_dates_from_meta(all_mentions, max_fetches=25)

    total = len(etransfer_social) + len(etransfer_press) + len(competitor_mentions)
    global LAST_FETCH_DIAGNOSTICS
    LAST_FETCH_DIAGNOSTICS = {
        "reddit_et": {
            "tasks": reddit_et_tasks_run,
            "exceptions": reddit_et_exceptions,
            "post_rows_in_batches": reddit_et_post_rows,
            "posts_after_reddit_et": after_reddit_et_only,
        },
        "reddit_comp": {
            "tasks": reddit_comp_tasks_run,
            "exceptions": reddit_comp_exceptions,
            "post_rows_in_batches": reddit_comp_post_rows,
            "posts_after_reddit_comp": after_reddit_comp_only,
        },
        "selenium": {
            "et_queries": selenium_et_query_count,
            "et_rows_returned": selenium_et_rows,
            "comp_queries": selenium_comp_query_count,
            "comp_rows_returned": selenium_comp_rows,
            "browser_concurrency_cap": SCRAPE_MAX_CONCURRENT_BROWSERS,
        },
        "ddg_fallback": {
            "used": ddg_fallback_used,
            "rows_added": ddg_rows_added,
            "enabled": FETCH_FALLBACK_DDG,
        },
        "totals": {
            "before_ddg": total_before_ddg,
            "final": total,
            "final_social": len(etransfer_social),
            "final_press": len(etransfer_press),
            "final_competitor": len(competitor_mentions),
        },
    }
    if total == 0:
        diag_lines = [
            "=== FETCH DIAGNOSTICS ===",
            f"Reddit JSON e-Transfer: {reddit_et_tasks_run} tasks, {reddit_et_exceptions} exceptions, "
            f"{reddit_et_post_rows} raw rows in responses, {after_reddit_et_only} posts after Reddit phase.",
            f"Reddit JSON competitor: {reddit_comp_tasks_run} tasks, {reddit_comp_exceptions} exceptions, "
            f"{reddit_comp_post_rows} raw rows, {after_reddit_comp_only} posts after Reddit competitor phase.",
            f"Selenium: {selenium_et_query_count} e-Transfer query bundles ({selenium_et_rows} rows), "
            f"{selenium_comp_query_count} competitor bundles ({selenium_comp_rows} rows); "
            f"max {SCRAPE_MAX_CONCURRENT_BROWSERS} concurrent browsers.",
            f"DDG fallback: {'ran' if ddg_fallback_used else 'skipped'} (FETCH_FALLBACK_DDG={FETCH_FALLBACK_DDG}); "
            f"rows added from DDG pass: {ddg_rows_added}.",
            f"Final buckets: social={len(etransfer_social)} press={len(etransfer_press)} competitor={len(competitor_mentions)}.",
            NO_MENTIONS_MARKER,
        ]
        return "\n".join(diag_lines)

    lines = [f"=== INTERAC BIWEEKLY SCAN — {now_est()} ==="]
    lines.append(
        f"Total: {total} mentions "
        f"({len(etransfer_social)} e-Transfer community, {len(etransfer_press)} e-Transfer news, "
        f"{len(competitor_mentions)} competitor)"
    )
    lines.append("")

    def _fmt(prefix: str, items: list[dict], cap: int, snippet_cap: int) -> list[str]:
        out = []
        for i, m in enumerate(items[:cap], 1):
            snippet = " ".join((m.get("snippet", "") or "").split())[:snippet_cap]
            out.append(f"[{prefix}{i}] {m.get('source', 'unknown')}")
            # Explicit Date: field on its own line so the LLM reliably picks it up
            out.append(f"  Date: {m['date']}" if m.get("date") else "  Date: unknown")
            out.append(f"  Title: {m.get('title', '')[:120]}")
            out.append(f"  Snippet: {snippet}")
            out.append(f"  URL: {m.get('link', '')}")
            out.append("")
        return out

    if etransfer_social:
        lines.append("=== e-TRANSFER COMMUNITY (REDDIT, RFD, X) ===")
        lines += _fmt("S", etransfer_social, 30, 500)

    if etransfer_press:
        lines.append("=== e-TRANSFER NEWS ===")
        lines += _fmt("EN", etransfer_press, 8, 300)

    if competitor_mentions:
        lines.append("=== COMPETITOR INTELLIGENCE (Wise, PayPal, Apple Pay, Wealthsimple, KOHO, Venmo, Zelle, Revolut, Neo, ACH, others) ===")
        lines += _fmt("C", competitor_mentions, 30, 450)

    return "\n".join(lines)


# ─── Kimi K2.5 Analysis ──────────────────────────────────────────────────────
async def call_kimi(system_prompt: str, user_content: str) -> str:
    # 8192 token limit total. System prompt ~500 tokens, output ~800 tokens.
    # Budget ~6000 tokens (~18k chars) for user content. Cap at 15k for safety.
    if len(user_content) > 15000:
        user_content = user_content[:15000] + "\n\n[... truncated]"

    async with httpx.AsyncClient(timeout=90) as client:
        response = await client.post(
            KIMI_API_URL,
            headers={
                "Authorization": f"Bearer {KIMI_API_KEY}",
                "Content-Type": "application/json",
            },
            json={
                "model": KIMI_MODEL,
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_content},
                ],
                "temperature": 0.3,
                "max_tokens": 2000,
            },
        )
        if response.status_code != 200:
            body = response.text
            logger.error(f"Kimi API {response.status_code}: {body}")
            raise Exception(f"Kimi API {response.status_code}: {body[:300]}")
        data = response.json()
        return data["choices"][0]["message"]["content"]


async def analyze_biweekly(mentions_text: str) -> str:
    """Run the universal biweekly scan analysis, inject prior memory for trend tracking."""
    config = load_prompts()
    prompt = config["biweekly_prompt"].replace("{timestamp}", now_est())

    # Inject previous scan context so the LLM can populate Trend vs Last Scan
    prev_memory = _load_biweekly_memory()
    user_content = mentions_text
    if prev_memory.get("last_scan_date"):
        etransfer_themes = "; ".join(prev_memory.get("etransfer_themes", [])) or "none on record"
        competitor_themes = "; ".join(prev_memory.get("competitor_themes", [])) or "none on record"
        user_content += (
            f"\n\n--- PREVIOUS SCAN CONTEXT (for Trend vs Last Scan section) ---\n"
            f"Last scan date: {prev_memory['last_scan_date']}\n"
            f"e-Transfer themes from last scan: {etransfer_themes}\n"
            f"Competitor themes from last scan: {competitor_themes}\n"
        )

    report = await call_kimi(prompt, user_content)

    # Persist memory and Excel after a successful analysis
    scan_date = now_est()
    themes = _extract_biweekly_themes(report)
    _save_biweekly_memory(themes, scan_date)
    _append_biweekly_excel(scan_date, report)

    return report


async def run_biweekly_scan(update: Update) -> None:
    """Run the universal biweekly e-Transfer intelligence scan and deliver to Telegram."""
    global last_report, last_mentions_raw
    tracked = _track_current_task()
    try:
        await update.message.reply_text(
            "Running biweekly e-Transfer intelligence scan (Reddit, X, RedFlagDeals, news)..."
        )
        try:
            mentions = await asyncio.wait_for(
                fetch_biweekly_mentions(), timeout=BIWEEKLY_FETCH_TIMEOUT
            )
        except asyncio.TimeoutError:
            await update.message.reply_text(
                f"⏱️ Fetch timed out ({BIWEEKLY_FETCH_TIMEOUT}s). Set BIWEEKLY_FETCH_TIMEOUT if needed."
            )
            return
        last_mentions_raw = mentions

        if NO_MENTIONS_MARKER in mentions:
            await update.message.reply_text(f"No data found this scan.\n\n{mentions[:3800]}")
            return

        await update.message.reply_text("Mentions collected. Analyzing with Kimi...")
        try:
            report = await asyncio.wait_for(
                analyze_biweekly(mentions), timeout=BIWEEKLY_ANALYZE_TIMEOUT
            )
        except asyncio.TimeoutError:
            await update.message.reply_text(
                f"⏱️ Analysis timed out ({BIWEEKLY_ANALYZE_TIMEOUT}s). Set BIWEEKLY_ANALYZE_TIMEOUT if needed."
            )
            return
        last_report = report

        await send_chunked_message(
            update,
            f"Interac e-Transfer Intelligence — {now_est()}\n\n{report}",
        )
    finally:
        _untrack_task(tracked)


def parse_email_modes() -> set[str]:
    # Supports: alert, weekly, always, comma-separated combinations.
    modes = {m.strip().lower() for m in EMAIL_SEND_MODE.split(",") if m.strip()}
    if not modes:
        modes = {"alert"}
    if "always" in modes:
        modes.update({"alert", "weekly"})
    return modes


WEEKDAY_TO_INDEX = {
    "monday": 0,
    "tuesday": 1,
    "wednesday": 2,
    "thursday": 3,
    "friday": 4,
    "saturday": 5,
    "sunday": 6,
}


def weekly_key(now_local: datetime) -> str:
    year, week_num, _ = now_local.isocalendar()
    return f"{year}-W{week_num}-{EMAIL_WEEKLY_DAY}-{EMAIL_WEEKLY_HOUR}"


def _should_send_email(
    *,
    trigger: str,
    now_local: datetime | None = None,
) -> tuple[bool, str]:
    """
    Decide whether we should send an email for this scan.
    Returns (should_send, reason).
    """
    if not EMAIL_ENABLED:
        return False, "EMAIL_ENABLED=0"

    modes = parse_email_modes()

    if trigger == "weekly" and "weekly" not in modes:
        return False, f"mode excludes weekly ({EMAIL_SEND_MODE})"

    if trigger == "weekly" and EMAIL_ALERT_DEDUP and now_local is not None:
        current_weekly_key = weekly_key(now_local)
        if current_weekly_key == last_weekly_email_key:
            return False, "weekly dedup"

    if EMAIL_COOLDOWN_MINUTES > 0 and last_email_sent_at is not None:
        minutes_since = (datetime.now(timezone.utc) - last_email_sent_at).total_seconds() / 60.0
        if minutes_since < EMAIL_COOLDOWN_MINUTES:
            return False, f"cooldown {minutes_since:.1f}m/{EMAIL_COOLDOWN_MINUTES}m"

    return True, "ok"


def _smtp_config_summary() -> str:
    recipient_count = len(EMAIL_TO)
    user_hint = SMTP_USERNAME if SMTP_USERNAME else "(empty)"
    return (
        f"host={SMTP_HOST or '(empty)'} port={SMTP_PORT} "
        f"user={user_hint} from={EMAIL_FROM or '(empty)'} recipients={recipient_count}"
    )


def _resend_config_summary() -> str:
    key_hint = "(set)" if RESEND_API_KEY else "(empty)"
    recipient_count = len(EMAIL_TO)
    return (
        f"url={RESEND_API_URL} key={key_hint} from={EMAIL_FROM or '(empty)'} "
        f"recipients={recipient_count}"
    )


def _validate_smtp_config() -> tuple[bool, str]:
    missing = []
    if not EMAIL_ENABLED:
        missing.append("EMAIL_ENABLED")
    if not SMTP_HOST:
        missing.append("SMTP_HOST")
    if not EMAIL_FROM:
        missing.append("EMAIL_FROM")
    if not EMAIL_TO:
        missing.append("EMAIL_TO")
    if not SMTP_USERNAME:
        missing.append("SMTP_USERNAME")
    if not SMTP_PASSWORD:
        missing.append("SMTP_PASSWORD")

    if missing:
        return False, f"Missing/invalid env vars: {', '.join(missing)}"
    return True, "ok"


def _validate_resend_config() -> tuple[bool, str]:
    missing = []
    if not EMAIL_ENABLED:
        missing.append("EMAIL_ENABLED")
    if not RESEND_API_KEY:
        missing.append("RESEND_API_KEY")
    if not EMAIL_FROM:
        missing.append("EMAIL_FROM")
    if not EMAIL_TO:
        missing.append("EMAIL_TO")
    if missing:
        return False, f"Missing/invalid env vars: {', '.join(missing)}"
    return True, "ok"


def _send_email_smtp(subject: str, body: str) -> tuple[bool, str]:
    valid, reason = _validate_smtp_config()
    if not valid:
        logger.warning(f"Email send skipped: {reason}")
        return False, reason

    try:
        text_body, html_body = build_email_bodies(subject, body)
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = EMAIL_FROM
        msg["To"] = ", ".join(EMAIL_TO)
        msg.attach(MIMEText(text_body, "plain", _charset="utf-8"))
        msg.attach(MIMEText(html_body, "html", _charset="utf-8"))

        if SMTP_PORT == 465:
            server = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=30)
        else:
            server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30)
            server.ehlo()
            server.starttls()
            server.ehlo()
        server.login(SMTP_USERNAME, SMTP_PASSWORD)
        server.sendmail(EMAIL_FROM, EMAIL_TO, msg.as_string())
        server.quit()
        return True, "email accepted by SMTP server"
    except Exception as e:
        logger.error(f"Failed to send email via SMTP: {e}")
        return False, str(e)


def _smtp_login_check() -> tuple[bool, str]:
    valid, reason = _validate_smtp_config()
    if not valid:
        return False, reason
    try:
        if SMTP_PORT == 465:
            server = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=30)
        else:
            server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30)
            server.ehlo()
            server.starttls()
            server.ehlo()
        server.login(SMTP_USERNAME, SMTP_PASSWORD)
        server.noop()
        server.quit()
        return True, "ok"
    except Exception as e:
        return False, str(e)


def _send_email_resend(subject: str, body: str) -> tuple[bool, str]:
    valid, reason = _validate_resend_config()
    if not valid:
        logger.warning(f"Email send skipped: {reason}")
        return False, reason

    try:
        text_body, html_body = build_email_bodies(subject, body)
        response = httpx.post(
            RESEND_API_URL,
            headers={
                "Authorization": f"Bearer {RESEND_API_KEY}",
                "Content-Type": "application/json",
            },
            json={
                "from": EMAIL_FROM,
                "to": EMAIL_TO,
                "subject": subject,
                "text": text_body,
                "html": html_body,
            },
            timeout=30,
        )
        if response.status_code not in (200, 201, 202):
            return False, f"Resend API {response.status_code}: {response.text[:300]}"
        return True, "email accepted by Resend API"
    except Exception as e:
        logger.error(f"Failed to send email via Resend: {e}")
        return False, str(e)


def smtp_health_check() -> tuple[bool, str]:
    if EMAIL_PROVIDER == "resend":
        valid, reason = _validate_resend_config()
        if not valid:
            return False, f"{reason}. Current: {_resend_config_summary()}"
        try:
            # Check API reachability + key validity via a lightweight domains call.
            response = httpx.get(
                "https://api.resend.com/domains",
                headers={"Authorization": f"Bearer {RESEND_API_KEY}"},
                timeout=30,
            )
            if response.status_code != 200:
                return False, f"Resend health check {response.status_code}: {response.text[:300]}"
            return True, f"Resend API reachable and key accepted. {_resend_config_summary()}"
        except Exception as e:
            return False, f"Resend health check failed: {e}. {_resend_config_summary()}"

    valid, reason = _validate_smtp_config()
    if not valid:
        return False, f"{reason}. Current: {_smtp_config_summary()}"
    ok, send_reason = _smtp_login_check()
    if ok:
        return True, f"SMTP connection/login successful. {_smtp_config_summary()}"
    return False, f"SMTP health check failed: {send_reason}. {_smtp_config_summary()}"


def send_email(subject: str, body: str) -> tuple[bool, str]:
    if EMAIL_PROVIDER == "resend":
        return _send_email_resend(subject, body)
    return _send_email_smtp(subject, body)


def _extract_report_field(report: str, field_name: str) -> str:
    pattern = rf"^{re.escape(field_name)}\s*:\s*(.+)$"
    m = re.search(pattern, report, flags=re.IGNORECASE | re.MULTILINE)
    return m.group(1).strip() if m else "N/A"


def _extract_section(report: str, start_marker: str, end_markers: list[str]) -> str:
    start_idx = report.find(start_marker)
    if start_idx == -1:
        return ""
    start_idx += len(start_marker)

    end_idx = len(report)
    for marker in end_markers:
        idx = report.find(marker, start_idx)
        if idx != -1:
            end_idx = min(end_idx, idx)
    return report[start_idx:end_idx].strip()


def _short_link_label(label: str, url: str) -> str:
    label = (label or "").strip()
    generic = {"source", "source url", "url", "link", "sourceurl"}
    if not label or label.lower().replace(" ", "") in generic:
        host = urlparse(url).netloc.replace("www.", "")
        label = host or "source"
    return label[:28] + "..." if len(label) > 31 else label


def _compact_email_line(raw_line: str) -> str:
    line = raw_line.strip()
    if not line:
        return ""

    line = re.sub(r"^\d+\.\s*", "", line)
    line = line.lstrip("- ").strip()
    line = line.replace("**", "").replace("`", "")

    links: list[str] = []

    def _store_link(label: str, url: str) -> str:
        safe_url = html.escape(url, quote=True)
        safe_label = html.escape(_short_link_label(label, url))
        links.append(
            f"<a href=\"{safe_url}\" style=\"font-size:13px;color:#175CD3;text-decoration:none;\">{safe_label}</a>"
        )
        return ""

    # Convert markdown links and remove them from body text.
    line = re.sub(
        r"\[([^\]]+)\]\((https?://[^\s)]+)\)",
        lambda m: _store_link(m.group(1), m.group(2)),
        line,
    )
    # Convert bare URLs into short domain links and remove from body text.
    line = re.sub(
        r"https?://[^\s)]+",
        lambda m: _store_link("", m.group(0)),
        line,
    )

    # If historical fields exist, collapse into a concise sentence.
    date_match = re.search(r"\bDate\s*:\s*([0-9]{4}-[0-9]{2}-[0-9]{2})", line, flags=re.IGNORECASE)
    product_match = re.search(r"\bProduct\s*:\s*([^:]+?)(?=\s+[A-Za-z ]+:\s|$)", line, flags=re.IGNORECASE)
    sentiment_match = re.search(r"\bSentiment Summary\s*:\s*(.+)$", line, flags=re.IGNORECASE)
    if sentiment_match:
        main_text = sentiment_match.group(1).strip()
        meta = []
        if product_match:
            meta.append(product_match.group(1).strip())
        if date_match:
            meta.append(date_match.group(1).strip())
        if meta:
            main_text = f"{main_text} ({', '.join(meta)})"
    else:
        # Keep explicit date labels in rendered bullets for faster evidence-time checks.
        line = re.sub(r"\b(Source URL|Product|Sentiment Summary)\s*:\s*", "", line, flags=re.IGNORECASE)
        main_text = " ".join(line.split())

    main_text = " ".join(main_text.split())
    if len(main_text) > 240:
        main_text = main_text[:237].rstrip() + "..."

    links_html = ""
    if links:
        links_html = " <span style=\"white-space:nowrap;\">" + " · ".join(links[:2]) + "</span>"
    if not main_text:
        return links_html
    return f"{html.escape(main_text)}{links_html}"


EMAIL_FONT_STACK = "Inter, -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif"
EMAIL_PAGE_BG = "#eef2f8"
EMAIL_CARD_BG = "#ffffff"
EMAIL_BORDER = "#d9e2f2"
EMAIL_NAVY = "#0f1f47"
EMAIL_TEXT = "#101828"
EMAIL_MUTED = "#667085"
EMAIL_ACCENT = "#175CD3"
EMAIL_CONTAINER_WIDTH = "920"
BIWEEKLY_MEMORY_PATH = Path(__file__).parent / "state" / "biweekly_memory.json"
BIWEEKLY_EXCEL_PATH = Path(__file__).parent / "state" / "biweekly_reports.xlsx"


def _compact_panel(raw: str, empty_msg: str, *, max_lines: int = 3, list_mode: bool = True) -> str:
    if not raw:
        return (
            f"<div style='border:1px solid {EMAIL_BORDER};border-radius:10px;padding:10px 12px;margin-top:6px;"
            f"font-size:13px;line-height:1.45;color:{EMAIL_MUTED};background:#fcfdff;'>{html.escape(empty_msg)}</div>"
        )
    lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
    items = []
    for ln in lines[:max_lines]:
        compact = _compact_email_line(ln)
        if compact:
            items.append(compact)
    if not items:
        return (
            f"<div style='border:1px solid {EMAIL_BORDER};border-radius:10px;padding:10px 12px;margin-top:6px;"
            f"font-size:13px;line-height:1.45;color:{EMAIL_MUTED};background:#fcfdff;'>{html.escape(empty_msg)}</div>"
        )
    if list_mode:
        joined = "".join(f"<li style='margin:0 0 4px 0;'>{it}</li>" for it in items)
        return (
            f"<div style='border:1px solid {EMAIL_BORDER};border-radius:10px;padding:10px 12px;margin-top:6px;background:#fcfdff;'>"
            f"<ul style='margin:0 0 0 16px;padding:0;color:{EMAIL_TEXT};font-size:13px;line-height:1.45;'>{joined}</ul>"
            "</div>"
        )
    joined = "<br>".join(items)
    return (
        f"<div style='border:1px solid {EMAIL_BORDER};border-radius:10px;padding:10px 12px;margin-top:6px;"
        f"font-size:13px;line-height:1.45;color:{EMAIL_TEXT};background:#fcfdff;'>{joined}</div>"
    )


def _load_biweekly_memory() -> dict:
    try:
        if not BIWEEKLY_MEMORY_PATH.exists():
            return {}
        return json.loads(BIWEEKLY_MEMORY_PATH.read_text())
    except Exception:
        return {}


def _save_biweekly_memory(themes: dict, scan_date: str) -> None:
    try:
        BIWEEKLY_MEMORY_PATH.parent.mkdir(parents=True, exist_ok=True)
        data = {
            "last_scan_date": scan_date,
            "etransfer_themes": themes.get("etransfer_themes", []),
            "competitor_themes": themes.get("competitor_themes", []),
        }
        BIWEEKLY_MEMORY_PATH.write_text(json.dumps(data, indent=2))
    except Exception as e:
        logger.warning(f"Could not save biweekly memory: {e}")


def _extract_biweekly_themes(report: str) -> dict:
    """Extract short theme labels from biweekly report sections for memory storage."""
    etransfer_raw = _extract_section(report, "e-Transfer Chatter:", ["Competitor Landscape:", "Trend vs Last Scan:"])
    competitor_raw = _extract_section(report, "Competitor Landscape:", ["Trend vs Last Scan:"])

    def _bullets_to_themes(section_text: str) -> list[str]:
        themes = []
        for line in (section_text or "").splitlines():
            stripped = line.strip()
            if stripped.startswith("- ") and "Nothing notable" not in stripped:
                quote_only = stripped[2:].strip()
                if " — " in quote_only:
                    quote_only = quote_only.split(" — ")[0].strip()
                theme = quote_only[:60]
                if theme:
                    themes.append(theme)
        return themes[:6]

    return {
        "etransfer_themes": _bullets_to_themes(etransfer_raw),
        "competitor_themes": _bullets_to_themes(competitor_raw),
    }


def _append_biweekly_excel(scan_date: str, report: str) -> None:
    """Append biweekly report sections to Excel file for human review."""
    try:
        from openpyxl import Workbook, load_workbook

        etransfer_raw = _extract_section(report, "e-Transfer Chatter:", ["Competitor Landscape:", "Trend vs Last Scan:"])
        competitor_raw = _extract_section(report, "Competitor Landscape:", ["Trend vs Last Scan:"])
        trend_raw = _extract_section(report, "Trend vs Last Scan:", [])

        BIWEEKLY_EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)

        if BIWEEKLY_EXCEL_PATH.exists():
            wb = load_workbook(BIWEEKLY_EXCEL_PATH)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Biweekly Reports"
            ws.append(["Scan Date", "e-Transfer Chatter", "Competitor Landscape", "Trend vs Last Scan", "Full Report"])

        ws.append([
            scan_date,
            (etransfer_raw or "").strip(),
            (competitor_raw or "").strip(),
            (trend_raw or "").strip(),
            report.strip(),
        ])
        wb.save(BIWEEKLY_EXCEL_PATH)
        logger.info(f"Appended biweekly report to {BIWEEKLY_EXCEL_PATH}")
    except Exception as e:
        logger.warning(f"Could not append to Excel: {e}")


def _styled_raw_report_html(subject: str, body: str) -> str:
    escaped = html.escape(body)
    return f"""
<html>
  <body style="margin:0;padding:0;background:{EMAIL_PAGE_BG};font-family:{EMAIL_FONT_STACK};color:{EMAIL_TEXT};">
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="padding:24px 0;">
      <tr><td align="center">
        <table role="presentation" width="{EMAIL_CONTAINER_WIDTH}" cellspacing="0" cellpadding="0" style="background:{EMAIL_CARD_BG};border-radius:14px;overflow:hidden;border:1px solid {EMAIL_BORDER};">
          <tr>
            <td style="background:{EMAIL_NAVY};color:#ffffff;padding:18px 24px;border-bottom:4px solid #fdb913;">
              <div style="font-size:24px;font-weight:700;letter-spacing:0.2px;">Interac Intelligence</div>
              <div style="font-size:13px;color:#d8def0;margin-top:6px;">{html.escape(subject)}</div>
            </td>
          </tr>
          <tr>
            <td style="padding:20px 24px;">
              <div style="font-size:15px;font-weight:700;margin-bottom:10px;">Report</div>
              <pre style="white-space:pre-wrap;background:#f9fbff;border:1px solid {EMAIL_BORDER};border-radius:10px;padding:14px;font-size:13px;line-height:1.5;color:{EMAIL_TEXT};">{escaped}</pre>
            </td>
          </tr>
        </table>
      </td></tr>
    </table>
  </body>
</html>
""".strip()


def _platform_badge_color(platform: str) -> str:
    """Return badge color for a platform name."""
    p = platform.lower()
    if "reddit" in p:
        return "#FF4500"
    if "redflagdeal" in p or "rfd" in p:
        return "#CC1200"
    if "twitter" in p or "x/" in p or p == "x":
        return "#1a1a1a"
    if "forum" in p:
        return "#6b7280"
    return "#1a73e8"  # News/Other


def _render_quote_bullets(raw: str, empty_msg: str) -> str:
    """Render quote bullets as styled cards with platform badge, date, and hyperlink.

    Each bullet is expected in the form:
        - "quote text" — Platform, Date. Source: URL
    """
    _empty = (
        f"<div style='padding:12px 14px;font-size:13px;color:{EMAIL_MUTED};'>"
        f"{html.escape(empty_msg)}</div>"
    )
    if not raw:
        return _empty

    items = []
    for line in raw.splitlines():
        stripped = line.strip()
        if not stripped or not (stripped.startswith("- ") or stripped.startswith("• ")):
            continue
        text = stripped[2:].strip()
        if not text or "Nothing notable" in text:
            continue

        # Split quote from attribution on ' — '
        if " — " in text:
            quote_part, attr_part = text.split(" — ", 1)
        else:
            quote_part, attr_part = text, ""

        # Extract URL — try "Source: URL" first, then bare https?:// in attr
        url = ""
        url_match = re.search(r"Source:\s*(https?://\S+)", attr_part, re.IGNORECASE)
        if url_match:
            url = url_match.group(1).rstrip(".,)")
            attr_part = re.sub(r"\s*Source:\s*https?://\S+", "", attr_part, flags=re.IGNORECASE).strip()
        else:
            bare_match = re.search(r"(https?://\S+)", attr_part)
            if bare_match:
                url = bare_match.group(1).rstrip(".,)")
                attr_part = re.sub(r"https?://\S+", "", attr_part).strip()

        # Parse date from attribution (everything after first comma).
        # Format with date:    — Platform, Date.
        # Format without date: — Platform.  (no comma, so no date)
        attr_clean = attr_part.rstrip(".")
        if "," in attr_clean:
            _platform_from_llm, date_label = attr_clean.split(",", 1)
            date_label = date_label.strip().rstrip(".")
        else:
            # No comma → attribution is platform only, no date
            date_label = ""

        # Derive platform badge from URL domain — only community platforms get a badge.
        # This prevents corporate sites (wise.com, paypal.com) from showing as badges.
        COMMUNITY_SOURCES = {"Reddit", "X/Twitter", "RedFlagDeals", "Forum"}
        show_badge = False
        platform_label = ""
        if url:
            _, url_source = _classify_channel_and_source(url)
            if url_source in COMMUNITY_SOURCES:
                platform_label = url_source
                show_badge = True

        badge_color = _platform_badge_color(platform_label) if show_badge else "#d1d5db"

        # Build link HTML
        link_html = ""
        if url:
            safe_url = html.escape(url, quote=True)
            domain = re.sub(r"^www\.", "", re.sub(r"https?://", "", url).split("/")[0])
            link_html = (
                f"<a href='{safe_url}' style='font-size:11px;color:{EMAIL_ACCENT};"
                f"text-decoration:none;'>{html.escape(domain)}</a>"
            )

        quote_html = html.escape(quote_part.strip())
        date_html = html.escape(date_label) if date_label else ""

        # Build meta row — margin-right on each element (flex gap unreliable in webmail)
        meta_inner = ""
        if show_badge and platform_label:
            meta_inner += (
                f"<span style='background:{badge_color};color:#fff;font-size:10px;"
                f"font-weight:700;padding:2px 7px;border-radius:999px;"
                f"letter-spacing:0.3px;white-space:nowrap;margin-right:6px;display:inline-block;'>"
                f"{html.escape(platform_label)}</span>"
            )
        if date_html:
            meta_inner += (
                f"<span style='font-size:11px;color:{EMAIL_MUTED};margin-right:6px;'>{date_html}</span>"
            )
        if link_html:
            meta_inner += link_html

        meta_row = (
            f"<div style='margin-top:5px;line-height:1.8;'>{meta_inner}</div>"
            if meta_inner else ""
        )

        card = (
            f"<div style='border-left:3px solid {badge_color};padding:8px 0 8px 12px;"
            f"margin-bottom:14px;'>"
            f"<div style='font-size:13px;line-height:1.55;color:{EMAIL_TEXT};'>{quote_html}</div>"
            f"{meta_row}"
            f"</div>"
        )
        items.append(card)

    if not items:
        return _empty

    return "".join(items)


def _trend_mini_card(label: str, content: str, accent: str) -> str:
    """Render a single Trend sub-column card."""
    safe_content = html.escape(content.strip()) if content.strip() else "none identified"
    return (
        f"<td style='width:33%;vertical-align:top;padding:0 8px 0 0;'>"
        f"<div style='background:#eef2f8;border-radius:10px;padding:12px 14px;height:100%;box-sizing:border-box;'>"
        f"<div style='font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.6px;"
        f"color:{accent};margin-bottom:6px;'>{html.escape(label)}</div>"
        f"<div style='font-size:12px;line-height:1.55;color:{EMAIL_TEXT};'>{safe_content}</div>"
        f"</div></td>"
    )


def _parse_trend_fields(trend_raw: str) -> tuple[str, str, str]:
    """Extract Still active / Went quiet / New this scan values from trend section."""
    still, quiet, new = "", "", ""
    for line in (trend_raw or "").splitlines():
        l = line.strip()
        if l.lower().startswith("- still active:"):
            still = l.split(":", 1)[1].strip()
        elif l.lower().startswith("- went quiet:"):
            quiet = l.split(":", 1)[1].strip()
        elif l.lower().startswith("- new this scan:"):
            new = l.split(":", 1)[1].strip()
    return still, quiet, new


def _build_biweekly_html(subject: str, body: str) -> str:
    scan_date = _extract_report_field(body, "SCAN DATE")
    etransfer_raw = _extract_section(body, "e-Transfer Chatter:", ["Competitor Landscape:", "Trend vs Last Scan:"])
    competitor_raw = _extract_section(body, "Competitor Landscape:", ["Trend vs Last Scan:"])
    trend_raw = _extract_section(body, "Trend vs Last Scan:", [])

    if not any(s.strip() for s in [etransfer_raw, competitor_raw, trend_raw]):
        return _styled_raw_report_html(subject, body)

    etransfer_html = _render_quote_bullets(etransfer_raw, "Nothing notable this scan.")
    competitor_html = _render_quote_bullets(competitor_raw, "Nothing notable this scan.")

    still, quiet, new_themes = _parse_trend_fields(trend_raw)
    trend_html = (
        f"<table role='presentation' width='100%' cellspacing='0' cellpadding='0'><tr>"
        + _trend_mini_card("Still active", still, "#027A48")
        + _trend_mini_card("Went quiet", quiet, "#667085")
        + _trend_mini_card("New this scan", new_themes, "#175CD3")
        + "</tr></table>"
    )

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
  .qcard:hover {{ background:#f5f8ff !important; }}
  a:hover {{ text-decoration:underline !important; }}
</style>
</head>
<body style="margin:0;padding:0;background:{EMAIL_PAGE_BG};font-family:{EMAIL_FONT_STACK};color:{EMAIL_TEXT};">
  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="padding:24px 0;">
    <tr><td align="center">
      <table role="presentation" width="1200" cellspacing="0" cellpadding="0" style="background:{EMAIL_CARD_BG};border-radius:14px;overflow:hidden;border:1px solid {EMAIL_BORDER};">

        <!-- HEADER -->
        <tr>
          <td colspan="2" style="background:{EMAIL_NAVY};color:#ffffff;padding:20px 28px;border-bottom:4px solid #fdb913;">
            <div style="font-size:24px;font-weight:700;letter-spacing:0.2px;">Interac e-Transfer Intelligence</div>
            <div style="font-size:13px;color:#aebce2;margin-top:5px;">{html.escape(scan_date)}</div>
          </td>
        </tr>

        <!-- TWO-COLUMN BODY -->
        <tr>
          <!-- LEFT: e-Transfer Chatter -->
          <td width="50%" style="vertical-align:top;padding:22px 14px 22px 28px;border-right:1px solid {EMAIL_BORDER};">
            <div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.6px;color:#c4320a;margin-bottom:4px;">Pain points</div>
            <div style="font-size:16px;font-weight:700;color:{EMAIL_TEXT};margin-bottom:16px;">e-Transfer Chatter</div>
            {etransfer_html}
          </td>
          <!-- RIGHT: Competitor Landscape -->
          <td width="50%" style="vertical-align:top;padding:22px 28px 22px 14px;">
            <div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.6px;color:#027A48;margin-bottom:4px;">What's working</div>
            <div style="font-size:16px;font-weight:700;color:{EMAIL_TEXT};margin-bottom:16px;">Competitor Landscape</div>
            {competitor_html}
          </td>
        </tr>

        <!-- TREND SECTION -->
        <tr>
          <td colspan="2" style="padding:0 28px 28px 28px;border-top:1px solid {EMAIL_BORDER};">
            <div style="font-size:16px;font-weight:700;color:{EMAIL_TEXT};margin:20px 0 12px 0;">Trend vs Last Scan</div>
            {trend_html}
          </td>
        </tr>

      </table>
    </td></tr>
  </table>
</body>
</html>""".strip()


def build_email_bodies(subject: str, body: str) -> tuple[str, str]:
    return body, _build_biweekly_html(subject, body)


def _record_email_sent(trigger: str, *, now_local: datetime | None = None) -> None:
    global last_email_sent_at, last_weekly_email_key
    last_email_sent_at = datetime.now(timezone.utc)
    if trigger == "weekly" and now_local is not None:
        last_weekly_email_key = weekly_key(now_local)


def weekly_est_to_utc(day_name: str, hour_est: int) -> tuple[int, int]:
    base_day = WEEKDAY_TO_INDEX.get(day_name, 0)
    hour_utc = hour_est + 5  # EST -> UTC
    day_shift = 0
    if hour_utc >= 24:
        hour_utc -= 24
        day_shift = 1
    return (base_day + day_shift) % 7, hour_utc


async def ask_followup(question: str, report_context: str) -> str:
    config = load_prompts()
    return await call_kimi(
        config["followup_prompt"],
        f"Latest report:\n{report_context}\n\nRaw mentions:\n{last_mentions_raw[:3000]}\n\nQuestion: {question}",
    )


# Telegram hard-limits message length (4096 chars). Chunk long reports safely.
async def send_chunked_message(
    update: Update,
    text: str,
    *,
    parse_mode: str | None = None,
    chunk_size: int = 3900,
) -> None:
    async def _reply_with_fallback(message_text: str) -> None:
        if not parse_mode:
            await update.message.reply_text(message_text)
            return
        try:
            await update.message.reply_text(message_text, parse_mode=parse_mode)
        except BadRequest as e:
            # Model-generated text can contain invalid markdown entities.
            # Retry as plain text so delivery succeeds instead of failing the command.
            if "Can't parse entities" in str(e):
                await update.message.reply_text(message_text)
                return
            raise

    if len(text) <= chunk_size:
        await _reply_with_fallback(text)
        return

    chunks = []
    remaining = text
    while remaining:
        if len(remaining) <= chunk_size:
            chunks.append(remaining)
            break
        split_at = remaining.rfind("\n\n", 0, chunk_size)
        if split_at < 0:
            split_at = remaining.rfind("\n", 0, chunk_size)
        if split_at < 0:
            split_at = chunk_size
        chunks.append(remaining[:split_at].rstrip())
        remaining = remaining[split_at:].lstrip()

    total = len(chunks)
    for idx, chunk in enumerate(chunks, 1):
        prefix = f"({idx}/{total})\n" if total > 1 else ""
        await _reply_with_fallback(prefix + chunk)


async def send_chunked_to_chat(
    bot,
    chat_id: int,
    text: str,
    *,
    chunk_size: int = 3900,
) -> None:
    """Send long text as multiple Telegram messages (webhook / background jobs)."""
    if len(text) <= chunk_size:
        await bot.send_message(chat_id=chat_id, text=text)
        return
    chunks = []
    remaining = text
    while remaining:
        if len(remaining) <= chunk_size:
            chunks.append(remaining)
            break
        split_at = remaining.rfind("\n\n", 0, chunk_size)
        if split_at < 0:
            split_at = remaining.rfind("\n", 0, chunk_size)
        if split_at < 0:
            split_at = chunk_size
        chunks.append(remaining[:split_at].rstrip())
        remaining = remaining[split_at:].lstrip()
    total = len(chunks)
    for idx, chunk in enumerate(chunks, 1):
        prefix = f"({idx}/{total})\n" if total > 1 else ""
        await bot.send_message(chat_id=chat_id, text=prefix + chunk)


async def reddit_json_health_probe() -> tuple[int, str]:
    """GET one post from r/personalfinancecanada/new.json — status for /fetchdiag."""
    try:
        async with httpx.AsyncClient(timeout=12, headers=_REDDIT_HEADERS) as client:
            r = await client.get(
                "https://www.reddit.com/r/personalfinancecanada/new.json",
                params={"limit": 1},
            )
        if r.status_code == 200:
            return 200, "ok"
        return r.status_code, (r.text[:120] or "")
    except Exception as e:
        return -1, f"{type(e).__name__}: {e}"[:120]


async def _background_biweekly_scan(bot, chat_id: int) -> None:
    global last_report, last_mentions_raw
    try:
        await bot.send_message(
            chat_id=chat_id,
            text="Running biweekly e-Transfer intelligence scan (Reddit, X, RedFlagDeals, news)...",
        )
        mentions = await asyncio.wait_for(
            fetch_biweekly_mentions(), timeout=BIWEEKLY_FETCH_TIMEOUT
        )
        last_mentions_raw = mentions
        if NO_MENTIONS_MARKER in mentions:
            await bot.send_message(
                chat_id=chat_id,
                text=f"No data found this scan.\n\n{mentions[:3800]}",
            )
            return
        await bot.send_message(chat_id=chat_id, text="Mentions collected. Analyzing with Kimi...")
        report = await asyncio.wait_for(
            analyze_biweekly(mentions), timeout=BIWEEKLY_ANALYZE_TIMEOUT
        )
        last_report = report
        await send_chunked_to_chat(
            bot,
            chat_id,
            f"Interac e-Transfer Intelligence — {now_est()}\n\n{report}",
        )
    except asyncio.TimeoutError:
        await bot.send_message(
            chat_id=chat_id,
            text=f"Timed out (fetch {BIWEEKLY_FETCH_TIMEOUT}s / analyze {BIWEEKLY_ANALYZE_TIMEOUT}s).",
        )
    except Exception as e:
        logger.error(f"Background scan failed: {e}")
        await bot.send_message(chat_id=chat_id, text=f"Scan failed: {e}")


async def _background_biweekly_email(bot, chat_id: int) -> None:
    global last_report, last_mentions_raw
    try:
        mentions = await asyncio.wait_for(
            fetch_biweekly_mentions(), timeout=BIWEEKLY_FETCH_TIMEOUT
        )
        last_mentions_raw = mentions
        if NO_MENTIONS_MARKER in mentions:
            await bot.send_message(
                chat_id=chat_id,
                text=f"No data found this scan.\n\n{mentions[:3800]}",
            )
            return
        await bot.send_message(chat_id=chat_id, text="Mentions collected. Analyzing with Kimi...")
        report = await asyncio.wait_for(
            analyze_biweekly(mentions), timeout=BIWEEKLY_ANALYZE_TIMEOUT
        )
        last_report = report
        subject = f"{EMAIL_SUBJECT_PREFIX} — MANUAL REPORT"
        body = f"Interac e-Transfer Intelligence — {now_est()}\n\n{report}"
        ok, send_reason = send_email(subject, body)
        if ok:
            _record_email_sent("on_demand")
            await bot.send_message(chat_id=chat_id, text="Email sent successfully.")
        else:
            await bot.send_message(chat_id=chat_id, text=f"Email failed: {send_reason}")
    except asyncio.TimeoutError:
        await bot.send_message(
            chat_id=chat_id,
            text=f"Timed out (fetch {BIWEEKLY_FETCH_TIMEOUT}s / analyze {BIWEEKLY_ANALYZE_TIMEOUT}s).",
        )
    except Exception as e:
        logger.error(f"Background /email failed: {e}")
        await bot.send_message(chat_id=chat_id, text=f"/email failed: {e}")


# ─── Scheduled Broadcast ─────────────────────────────────────────────────────
async def scheduled_biweekly_broadcast(context: ContextTypes.DEFAULT_TYPE):
    """Daily job that runs the biweekly scan if 14+ days have passed since the last one."""
    global last_report, last_mentions_raw

    memory = _load_biweekly_memory()
    last_date_str = memory.get("last_scan_date")
    if last_date_str:
        try:
            last_date = datetime.fromisoformat(last_date_str)
            if not last_date.tzinfo:
                last_date = last_date.replace(tzinfo=EST)
            days_since = (datetime.now(EST) - last_date).days
            if days_since < 14:
                logger.info(f"Biweekly scan skipped: {days_since} days since last scan (need 14).")
                return
        except Exception as e:
            logger.warning(f"Could not parse last_scan_date for biweekly guard: {e}")

    tracked = _track_current_task()
    logger.info(f"[{now_est()}] Running scheduled biweekly scan...")
    try:
        try:
            mentions = await asyncio.wait_for(
                fetch_biweekly_mentions(), timeout=BIWEEKLY_FETCH_TIMEOUT
            )
        except asyncio.TimeoutError:
            logger.error(f"Scheduled biweekly: fetch timed out ({BIWEEKLY_FETCH_TIMEOUT}s)")
            return
        last_mentions_raw = mentions
        if NO_MENTIONS_MARKER in mentions:
            logger.warning("Scheduled biweekly: fetch returned no mentions; skipping Kimi and broadcast.")
            return
        try:
            report = await asyncio.wait_for(
                analyze_biweekly(mentions), timeout=BIWEEKLY_ANALYZE_TIMEOUT
            )
        except asyncio.TimeoutError:
            logger.error(f"Scheduled biweekly: analysis timed out ({BIWEEKLY_ANALYZE_TIMEOUT}s)")
            return
        last_report = report

        message = f"Interac e-Transfer Intelligence — {now_est()}\n\n{report}"
        for chat_id in subscribed_chats.copy():
            try:
                await context.bot.send_message(chat_id=chat_id, text=message)
            except Exception as e:
                logger.error(f"Failed to send biweekly report to {chat_id}: {e}")
                subscribed_chats.discard(chat_id)

        now_local = datetime.now(EST)
        should_send, reason = _should_send_email(trigger="weekly", now_local=now_local)
        if should_send:
            subject = f"{EMAIL_SUBJECT_PREFIX} — BIWEEKLY REPORT"
            ok, send_reason = send_email(subject, f"Interac e-Transfer Intelligence — {now_est()}\n\n{report}")
            if ok:
                _record_email_sent("weekly", now_local=now_local)
            else:
                logger.error(f"Biweekly email failed: {send_reason}")
        else:
            logger.info(f"Biweekly email not sent: {reason}")
    except Exception as e:
        logger.error(f"Scheduled biweekly scan failed: {e}")
    finally:
        _untrack_task(tracked)


# ─── Command Handlers ────────────────────────────────────────────────────────
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    subscribed_chats.add(update.effective_chat.id)
    await update.message.reply_text(
        "👋 *Interac Intelligence Bot*\n\n"
        "Scans Reddit, X, RedFlagDeals, and news for e-Transfer chatter. "
        "Biweekly report delivered automatically.\n\n"
        "*Commands:*\n"
        "• /subscribe — Get scheduled biweekly reports\n"
        "• /unsubscribe — Stop reports\n"
        "• /scan — Run biweekly scan now\n"
        "• /raw — See raw mentions from last scan\n"
        "• /prompt — View current config\n"
        "• /status — Check bot status\n"
        "• /email — Admin: run fresh biweekly scan + send email\n"
        "• /stop — Admin: cancel running jobs\n"
        "• /smtpcheck — Admin: check email config\n"
        "• /fetchdiag — Admin: Reddit + Selenium + DDG connectivity probe\n"
        "• Any text → Follow-up on latest report",
        parse_mode="Markdown",
    )


async def cmd_subscribe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    subscribed_chats.add(update.effective_chat.id)
    await update.message.reply_text("✅ Subscribed.")


async def cmd_unsubscribe(update: Update, context: ContextTypes.DEFAULT_TYPE):
    subscribed_chats.discard(update.effective_chat.id)
    await update.message.reply_text("🔕 Unsubscribed.")


async def cmd_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    config = load_prompts()
    et_count = len(config.get("etransfer_queries", config.get("biweekly_queries", [])))
    comp_count = len(config.get("competitor_queries", []))
    memory = _load_biweekly_memory()
    last_scan = memory.get("last_scan_date", "never")
    is_admin = "✅" if update.effective_user.id in ADMIN_IDS else "❌"
    await update.message.reply_text(
        f"✅ Bot running — {now_est()}\n"
        f"Selenium + Reddit | timeouts fetch {BIWEEKLY_FETCH_TIMEOUT}s / analyze {BIWEEKLY_ANALYZE_TIMEOUT}s\n"
        f"DDG fallback if empty: {FETCH_FALLBACK_DDG} | webhook: {bool(WEBHOOK_URL.strip())}\n"
        f"e-Transfer queries: {et_count} | Competitor queries: {comp_count}\n"
        f"Last biweekly scan: {last_scan}\n"
        f"Schedule: daily check at 9am EST, runs every 14 days\n"
        f"Subscribed chats: {len(subscribed_chats)}\n"
        f"Admin: {is_admin}\n"
        f"Your ID: `{update.effective_user.id}`"
    )


async def cmd_scan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if WEBHOOK_URL.strip():
            asyncio.create_task(
                _background_biweekly_scan(context.bot, update.effective_chat.id),
                name="biweekly_scan_bg",
            )
            await update.message.reply_text(
                "Scan started in background (webhook mode). You will get more messages here when done."
            )
            return
        await run_biweekly_scan(update)
    except Exception as e:
        logger.error(f"Scan failed: {e}")
        await update.message.reply_text(f"❌ Scan failed: {e}")


async def cmd_raw(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not last_mentions_raw:
        await update.message.reply_text("No scan data yet. Run /scan first.")
        return
    text = last_mentions_raw[:4000]
    await update.message.reply_text(f"```\n{text}\n```", parse_mode="Markdown")


async def cmd_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    config = load_prompts()
    et_queries = config.get("etransfer_queries", config.get("biweekly_queries", []))
    comp_queries = config.get("competitor_queries", [])
    sample_et = "\n".join(f"  • {q}" for q in et_queries[:3])
    sample_comp = "\n".join(f"  • {q}" for q in comp_queries[:3])
    await update.message.reply_text(
        f"*e-Transfer queries:* {len(et_queries)}\n{sample_et}\n\n"
        f"*Competitor queries:* {len(comp_queries)}\n{sample_comp}\n\n"
        f"Edit `prompts.json` to change queries.\n"
        f"Edit `prompts/biweekly_prompt.md` to change the report format.",
        parse_mode="Markdown",
    )


async def cmd_fetchdiag(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("Admin only.")
        return
    code, hint = await reddit_json_health_probe()
    lines = [
        f"Reddit probe (r/personalfinancecanada/new.json): HTTP {code} {hint}",
    ]
    try:
        news = await selenium_source_search("Interac e-Transfer Canada", "news", 2)
        lines.append(f"Selenium Google News sample: {len(news)} row(s)")
    except Exception as e:
        lines.append(f"Selenium news probe error: {type(e).__name__}: {e}")
    try:
        ddg = await web_search("Interac e-Transfer", "search", 2, tbs="qdr:m")
        lines.append(f"DDG text sample: {len(ddg)} row(s)")
    except Exception as e:
        lines.append(f"DDG probe error: {type(e).__name__}: {e}")
    lines.append(f"WEBHOOK_URL set: {bool(WEBHOOK_URL.strip())}")
    lines.append(f"FETCH_FALLBACK_DDG: {FETCH_FALLBACK_DDG}")
    lines.append(f"Selenium browser cap: {SCRAPE_MAX_CONCURRENT_BROWSERS}")
    await update.message.reply_text("\n".join(lines)[:4000])


async def cmd_email(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global last_report, last_mentions_raw
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Admin only.")
        return

    tracked = _track_current_task()
    try:
        if WEBHOOK_URL.strip():
            asyncio.create_task(
                _background_biweekly_email(context.bot, update.effective_chat.id),
                name="biweekly_email_bg",
            )
            await update.message.reply_text(
                "📧 Email job started in background (webhook mode). Watch this chat for status."
            )
            return
        await update.message.reply_text("📧 Running fresh biweekly scan and sending email...")
        try:
            mentions = await asyncio.wait_for(
                fetch_biweekly_mentions(), timeout=BIWEEKLY_FETCH_TIMEOUT
            )
        except asyncio.TimeoutError:
            await update.message.reply_text(
                f"⏱️ Fetch timed out ({BIWEEKLY_FETCH_TIMEOUT}s). Set BIWEEKLY_FETCH_TIMEOUT if needed."
            )
            return
        last_mentions_raw = mentions

        if NO_MENTIONS_MARKER in mentions:
            await update.message.reply_text(f"No data found this scan.\n\n{mentions[:3800]}")
            return

        await update.message.reply_text("Mentions collected. Analyzing with Kimi...")
        try:
            report = await asyncio.wait_for(
                analyze_biweekly(mentions), timeout=BIWEEKLY_ANALYZE_TIMEOUT
            )
        except asyncio.TimeoutError:
            await update.message.reply_text(
                f"⏱️ Analysis timed out ({BIWEEKLY_ANALYZE_TIMEOUT}s). Set BIWEEKLY_ANALYZE_TIMEOUT if needed."
            )
            return
        last_report = report

        subject = f"{EMAIL_SUBJECT_PREFIX} — MANUAL REPORT"
        body = f"Interac e-Transfer Intelligence — {now_est()}\n\n{report}"
        ok, send_reason = send_email(subject, body)
        if ok:
            _record_email_sent("on_demand")
            await update.message.reply_text("✅ Email sent successfully.")
        else:
            await update.message.reply_text(f"❌ Email failed: {send_reason}")
    except Exception as e:
        logger.error(f"/email failed: {e}")
        await update.message.reply_text(f"❌ /email failed: {e}")
    finally:
        _untrack_task(tracked)


async def cmd_stop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Admin only.")
        return

    current = asyncio.current_task()
    cancelled = _cancel_active_tasks(exclude=current)
    await update.message.reply_text(f"🛑 Stop requested. Cancelled {cancelled} running task(s).")


async def cmd_smtpcheck(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Admin only.")
        return

    ok, reason = smtp_health_check()
    if ok:
        await update.message.reply_text(f"✅ {reason}")
    else:
        await update.message.reply_text(f"❌ {reason}")


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_text = update.message.text
    if not user_text:
        return
    if not last_report:
        await update.message.reply_text("No report yet. Run /scan first.")
        return

    allowed, remaining = check_rate_limit(update.effective_user.id)
    if not allowed:
        await update.message.reply_text(
            f"⚠️ Daily limit reached ({DAILY_LIMIT} questions/day). Resets at midnight EST."
        )
        return

    await update.message.reply_text("🤔 Thinking...")
    try:
        response = await ask_followup(user_text, last_report)
        suffix = f"\n\n_({remaining} questions remaining today)_" if remaining >= 0 else ""
        await update.message.reply_text(response + suffix, parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    app = Application.builder().token(TELEGRAM_TOKEN).build()

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("help", cmd_start))
    app.add_handler(CommandHandler("subscribe", cmd_subscribe))
    app.add_handler(CommandHandler("unsubscribe", cmd_unsubscribe))
    app.add_handler(CommandHandler("status", cmd_status))
    app.add_handler(CommandHandler("scan", cmd_scan))
    app.add_handler(CommandHandler("raw", cmd_raw))
    app.add_handler(CommandHandler("prompt", cmd_prompt))
    app.add_handler(CommandHandler("email", cmd_email))
    app.add_handler(CommandHandler("stop", cmd_stop))
    app.add_handler(CommandHandler("smtpcheck", cmd_smtpcheck))
    app.add_handler(CommandHandler("fetchdiag", cmd_fetchdiag))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    # Biweekly scan: runs daily at 9am EST (14:00 UTC) but self-guards to only execute every 14 days.
    job_queue = app.job_queue
    job_queue.run_daily(
        scheduled_biweekly_broadcast,
        time=datetime.strptime("14:00", "%H:%M").time(),
        name="biweekly_scan",
    )

    if WEBHOOK_URL:
        app.run_webhook(
            listen="0.0.0.0",
            port=PORT,
            webhook_url=f"{WEBHOOK_URL}/webhook",
            url_path="webhook",
        )
    else:
        app.run_polling()


if __name__ == "__main__":
    main()
